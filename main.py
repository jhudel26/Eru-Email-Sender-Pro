import sys
import os
import time
import re
import json
import logging
import sqlite3
from datetime import datetime
import pandas as pd
import win32com.client
from openpyxl import Workbook
from openpyxl.styles import Font

from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QTableWidget, QTableWidgetItem,
    QProgressBar, QTextEdit, QMessageBox, QSplitter,
    QLineEdit, QToolBar, QLabel, QFrame, QScrollArea,
    QGroupBox, QSizePolicy, QSpacerItem, QHeaderView, QComboBox,
    QStackedWidget, QFormLayout, QCheckBox, QSpinBox, QGridLayout,
    QMenu, QToolButton, QWidgetAction,
)
from PySide6.QtGui import QFont, QAction, QIcon, QPalette, QColor, QPixmap, QTextCursor, QTextBlockFormat, QKeySequence
from PySide6.QtCore import Qt, QThread, Signal, QSize, QTimer

from ui.design_system import build_stylesheet, LIGHT, DARK, SPACING
from ui.components import (
    AppSidebar, PageHeader, StatCard, ContentScrollArea,
    ToastManager, StatusBadge, make_field_label, make_section,
)

# =====================================================
# DATABASE MANAGER
# =====================================================
class DatabaseManager:
    """Manages SQLite database for persistent storage"""
    
    def __init__(self, db_name="eru_email_sender.db"):
        # Handle both script and executable environments
        if getattr(sys, 'frozen', False):
            # Running as PyInstaller executable
            app_data_dir = os.path.join(os.path.expanduser("~"), "EruEmailSender")
            os.makedirs(app_data_dir, exist_ok=True)
            self.db_path = os.path.join(app_data_dir, db_name)
        else:
            # Running as script
            script_dir = os.path.dirname(os.path.abspath(__file__))
            self.db_path = os.path.join(script_dir, db_name)
        
        self.initialize_database()
    
    def initialize_database(self):
        """Create database tables if they don't exist"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Settings table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Templates table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS templates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    subject TEXT,
                    body TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Campaigns table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS campaigns (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    total_recipients INTEGER DEFAULT 0,
                    confirmed_count INTEGER DEFAULT 0,
                    failed_count INTEGER DEFAULT 0,
                    unknown_count INTEGER DEFAULT 0,
                    cancelled_count INTEGER DEFAULT 0,
                    status TEXT DEFAULT 'pending',
                    started_at TIMESTAMP,
                    completed_at TIMESTAMP,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Recipients table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS recipients (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    campaign_id TEXT NOT NULL,
                    account TEXT,
                    full_name TEXT,
                    email TEXT,
                    cc TEXT,
                    attachment_path TEXT,
                    status TEXT DEFAULT 'pending',
                    attempt_number INTEGER DEFAULT 0,
                    last_error TEXT,
                    last_attempt_time TIMESTAMP,
                    row_index INTEGER,
                    FOREIGN KEY (campaign_id) REFERENCES campaigns(id)
                )
            ''')
            
            # Send attempts table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS send_attempts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    recipient_id INTEGER NOT NULL,
                    campaign_id TEXT NOT NULL,
                    attempt_number INTEGER NOT NULL,
                    status TEXT NOT NULL,
                    error_message TEXT,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (recipient_id) REFERENCES recipients(id),
                    FOREIGN KEY (campaign_id) REFERENCES campaigns(id)
                )
            ''')
            
            # Send logs table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS send_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    campaign_id TEXT,
                    recipient_id INTEGER,
                    log_level TEXT DEFAULT 'info',
                    message TEXT,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (campaign_id) REFERENCES campaigns(id),
                    FOREIGN KEY (recipient_id) REFERENCES recipients(id)
                )
            ''')
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"Error initializing database: {e}")
    
    def get_connection(self):
        """Get a database connection"""
        return sqlite3.connect(self.db_path)
    
    def save_setting(self, key, value):
        """Save a setting to the database"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT OR REPLACE INTO settings (key, value, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            ''', (key, json.dumps(value)))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error saving setting: {e}")
            return False
    
    def get_setting(self, key, default=None):
        """Get a setting from the database"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT value FROM settings WHERE key = ?', (key,))
            result = cursor.fetchone()
            conn.close()
            
            if result:
                value = result[0]
                # Try to parse as JSON
                try:
                    return json.loads(value)
                except:
                    return value
            return default
        except Exception as e:
            print(f"Error getting setting: {e}")
            return default
    
    def create_campaign(self, campaign_id, name, total_recipients):
        """Create a new campaign"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO campaigns (id, name, total_recipients, status, started_at)
                VALUES (?, ?, ?, 'in_progress', CURRENT_TIMESTAMP)
            ''', (campaign_id, name, total_recipients))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error creating campaign: {e}")
            return False
    
    def update_campaign_status(self, campaign_id, status):
        """Update campaign status"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            if status == 'completed':
                cursor.execute('''
                    UPDATE campaigns 
                    SET status = ?, completed_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (status, campaign_id))
            else:
                cursor.execute('''
                    UPDATE campaigns SET status = ? WHERE id = ?
                ''', (status, campaign_id))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error updating campaign status: {e}")
            return False
    
    def update_campaign_counts(self, campaign_id):
        """Update campaign count statistics"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Status values are stored as EmailState titles (e.g. "Confirmed"), not lowercase.
            cursor.execute('''
                UPDATE campaigns SET
                    confirmed_count = (SELECT COUNT(*) FROM recipients WHERE campaign_id = ? AND lower(status) = 'confirmed'),
                    failed_count = (SELECT COUNT(*) FROM recipients WHERE campaign_id = ? AND lower(status) = 'failed'),
                    unknown_count = (SELECT COUNT(*) FROM recipients WHERE campaign_id = ? AND lower(status) = 'unknown'),
                    cancelled_count = (SELECT COUNT(*) FROM recipients WHERE campaign_id = ? AND lower(status) = 'cancelled')
                WHERE id = ?
            ''', (campaign_id, campaign_id, campaign_id, campaign_id, campaign_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error updating campaign counts: {e}")
            return False
    
    def add_recipient(self, campaign_id, account, full_name, email, cc, attachment_path, row_index):
        """Add a recipient to a campaign"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO recipients (campaign_id, account, full_name, email, cc, attachment_path, row_index)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (campaign_id, account, full_name, email, cc, attachment_path, row_index))
            conn.commit()
            recipient_id = cursor.lastrowid
            conn.close()
            return recipient_id
        except Exception as e:
            print(f"Error adding recipient: {e}")
            return None
    
    def update_recipient_status(self, recipient_id, status, error_message=None):
        """Update recipient status"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Increment attempt number
            cursor.execute('''
                UPDATE recipients 
                SET status = ?, 
                    attempt_number = attempt_number + 1,
                    last_error = ?,
                    last_attempt_time = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (status, error_message, recipient_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error updating recipient status: {e}")
            return False
    
    def log_send_attempt(self, recipient_id, campaign_id, attempt_number, status, error_message=None):
        """Log a send attempt"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO send_attempts (recipient_id, campaign_id, attempt_number, status, error_message)
                VALUES (?, ?, ?, ?, ?)
            ''', (recipient_id, campaign_id, attempt_number, status, error_message))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error logging send attempt: {e}")
            return False
    
    def log_message(self, campaign_id, recipient_id, log_level, message):
        """Log a message"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO send_logs (campaign_id, recipient_id, log_level, message)
                VALUES (?, ?, ?, ?)
            ''', (campaign_id, recipient_id, log_level, message))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error logging message: {e}")
            return False
    
    def get_interrupted_campaigns(self):
        """Get campaigns that were interrupted (in_progress status)"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, name, total_recipients, confirmed_count, failed_count, 
                       unknown_count, cancelled_count, started_at
                FROM campaigns 
                WHERE status = 'in_progress'
                ORDER BY started_at DESC
            ''')
            campaigns = cursor.fetchall()
            conn.close()
            return campaigns
        except Exception as e:
            print(f"Error getting interrupted campaigns: {e}")
            return []
    
    def get_campaign_recipients(self, campaign_id):
        """Get all recipients for a campaign"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, account, full_name, email, cc, attachment_path, status, 
                       attempt_number, last_error, row_index
                FROM recipients 
                WHERE campaign_id = ?
                ORDER BY row_index
            ''', (campaign_id,))
            recipients = cursor.fetchall()
            conn.close()
            return recipients
        except Exception as e:
            print(f"Error getting campaign recipients: {e}")
            return []
    
    def get_pending_recipients(self, campaign_id):
        """Get pending recipients for a campaign"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, account, full_name, email, cc, attachment_path, row_index
                FROM recipients 
                WHERE campaign_id = ? AND lower(status) = 'pending'
                ORDER BY row_index
            ''', (campaign_id,))
            recipients = cursor.fetchall()
            conn.close()
            return recipients
        except Exception as e:
            print(f"Error getting pending recipients: {e}")
            return []

# =====================================================
# SETTINGS MANAGER
# =====================================================
class SettingsManager:
    def __init__(self, config_file="settings.json"):
        # Initialize database manager
        self.db_manager = DatabaseManager()
        
        # Handle both script and executable environments for JSON fallback
        if getattr(sys, 'frozen', False):
            # Running as PyInstaller executable
            # Use user's home directory for settings to ensure writability
            app_data_dir = os.path.join(os.path.expanduser("~"), "EruEmailSender")
            os.makedirs(app_data_dir, exist_ok=True)
            self.config_file = os.path.join(app_data_dir, config_file)
        else:
            # Running as script
            script_dir = os.path.dirname(os.path.abspath(__file__))
            self.config_file = os.path.join(script_dir, config_file)
        
        self.default_settings = {
            "window_geometry": None,
            "paragraph_spacing": 12,
            "email_templates": {},
            "last_excel_path": "",
            "auto_save_interval": 5,
            "retry_failed_emails": True,
            "max_retries": 3,
            "last_selected_template": "default",
            "theme": "light",
            "importance": 2,
            "send_delay": 0,
            "request_read_receipt": True,
        }
        
        # Migrate from JSON to database if JSON exists
        self.migrate_from_json()
        
        # Load settings from database
        self.settings = self.load_settings()
    
    def migrate_from_json(self):
        """Migrate settings from JSON file to database"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    json_settings = json.load(f)
                    # Save each setting to database
                    for key, value in json_settings.items():
                        self.db_manager.save_setting(key, value)
                # Backup and remove old JSON file
                backup_file = self.config_file + ".backup"
                os.rename(self.config_file, backup_file)
                print(f"Migrated settings from JSON to database. Backup saved to {backup_file}")
            except Exception as e:
                print(f"Error migrating settings: {e}")
    
    def _coerce(self, key, value):
        """Convert stored strings (e.g. 'True', '3') to the default setting type."""
        default = self.default_settings.get(key)
        if value is None:
            return default
        if isinstance(default, bool):
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return bool(value)
            return str(value).strip().lower() in ("true", "1", "yes", "on")
        if isinstance(default, int) and not isinstance(default, bool):
            try:
                return int(value)
            except (TypeError, ValueError):
                return default
        if isinstance(default, dict):
            return value if isinstance(value, dict) else default
        if isinstance(default, list):
            return value if isinstance(value, list) else default
        return value

    def load_settings(self):
        """Load settings from database"""
        try:
            settings = self.default_settings.copy()
            for key in self.default_settings.keys():
                value = self.db_manager.get_setting(key)
                if value is not None:
                    settings[key] = self._coerce(key, value)
            return settings
        except Exception as e:
            print(f"Error loading settings: {e}")
            return self.default_settings.copy()
    
    def save_settings(self):
        """Save settings to database"""
        try:
            for key, value in self.settings.items():
                self.db_manager.save_setting(key, value)
            return True
        except Exception:
            return False
    
    def get(self, key, default=None):
        value = self.settings.get(key, default)
        if key in self.default_settings:
            return self._coerce(key, value if value is not None else default)
        return value
    
    def set(self, key, value):
        self.settings[key] = value
        self.db_manager.save_setting(key, value)

# =====================================================
# EMAIL STATE MACHINE
# =====================================================
class EmailState:
    """Email sending states for proper state tracking"""
    PENDING = "Pending"
    VALIDATING = "Validating"
    SENDING = "Sending"
    SUBMITTED = "Submitted"
    CONFIRMED = "Confirmed"
    FAILED = "Failed"
    UNKNOWN = "Unknown"
    SKIPPED = "Skipped"
    CANCELLED = "Cancelled"
    
    @staticmethod
    def is_terminal(state):
        """States that are final and should not be automatically retried"""
        return state in [EmailState.CONFIRMED, EmailState.SKIPPED, EmailState.CANCELLED]
    
    @staticmethod
    def can_retry(state):
        """States that can be safely retried"""
        return state in [EmailState.FAILED]
    
    @staticmethod
    def is_uncertain(state):
        """States where we're unsure if the email was sent"""
        return state in [EmailState.UNKNOWN, EmailState.SUBMITTED]

# =====================================================
# OUTLOOK CLIENT MANAGER
# =====================================================
class OutlookClient:
    """Manages Outlook COM connection lifecycle"""
    
    def __init__(self):
        self.outlook = None
        self.namespace = None
        self.outbox = None
        self.sent = None
        self.is_connected = False
    
    def connect(self, max_attempts=5, retry_delay=3):
        """Connect to Outlook with retry logic"""
        import pythoncom
        
        for attempt in range(max_attempts):
            try:
                pythoncom.CoInitialize()
                self.outlook = win32com.client.Dispatch("Outlook.Application")
                self.namespace = self.outlook.GetNamespace("MAPI")
                self.outbox = self.namespace.GetDefaultFolder(4)  # olFolderOutbox
                self.sent = self.namespace.GetDefaultFolder(5)    # olFolderSentMail
                self.is_connected = True
                return True, "Connected to Outlook successfully"
            except Exception as e:
                if attempt < max_attempts - 1:
                    time.sleep(retry_delay)
                else:
                    return False, f"Failed to connect to Outlook after {max_attempts} attempts: {str(e)}"
        
        return False, "Unknown connection error"
    
    def disconnect(self):
        """Safely disconnect from Outlook"""
        try:
            import pythoncom
            self.is_connected = False
            self.outlook = None
            self.namespace = None
            self.outbox = None
            self.sent = None
            pythoncom.CoUninitialize()
        except Exception:
            pass  # Best effort cleanup
    
    def is_available(self):
        """Check if Outlook connection is available"""
        if not self.is_connected or self.outlook is None:
            return False
        try:
            # Try to access a simple property to test connection
            _ = self.outlook.Version
            return True
        except Exception:
            self.is_connected = False
            return False
    
    def create_email(self):
        """Create a new email item"""
        if not self.is_available():
            raise Exception("Outlook is not available")
        return self.outlook.CreateItem(0)  # olMailItem

    def get_outbox_entry_ids(self):
        """Return EntryIDs currently in the Outbox. Outbox is small; Sent folder is not used."""
        ids = set()
        try:
            items = self.outbox.Items
            item = items.GetFirst()
            while item is not None:
                try:
                    ids.add(item.EntryID)
                except Exception:
                    pass
                item = items.GetNext()
        except Exception:
            pass
        return ids

    def trigger_send_receive(self):
        """Ask Outlook to flush the Outbox / sync mail."""
        try:
            self.namespace.SendAndReceive(False)
        except Exception:
            try:
                # Fallback: sync all accounts
                syncs = self.namespace.SyncObjects
                for i in range(1, syncs.Count + 1):
                    try:
                        syncs.Item(i).Start()
                    except Exception:
                        pass
            except Exception:
                pass

    def _recipient_addresses(self, mail_item):
        """Collect display/SMTP addresses from a mail item (handles Exchange X.500)."""
        addresses = set()
        try:
            to_field = str(getattr(mail_item, "To", "") or "")
            if to_field:
                addresses.add(to_field.lower())
        except Exception:
            pass

        try:
            recipients = mail_item.Recipients
            for i in range(1, recipients.Count + 1):
                recip = recipients.Item(i)
                try:
                    recip.Resolve()
                except Exception:
                    pass

                for attr in ("Address", "Name"):
                    try:
                        val = str(getattr(recip, attr, "") or "").strip()
                        if val:
                            addresses.add(val.lower())
                    except Exception:
                        pass

                # PR_SMTP_ADDRESS — real email even when Address is X.500
                try:
                    smtp = recip.PropertyAccessor.GetProperty(
                        "http://schemas.microsoft.com/mapi/proptag/0x39FE001E"
                    )
                    if smtp:
                        addresses.add(str(smtp).strip().lower())
                except Exception:
                    pass

                try:
                    entry = recip.AddressEntry
                    if entry is not None:
                        try:
                            user = entry.GetExchangeUser()
                            if user is not None and user.PrimarySmtpAddress:
                                addresses.add(str(user.PrimarySmtpAddress).strip().lower())
                        except Exception:
                            pass
                        try:
                            addr = str(getattr(entry, "Address", "") or "").strip()
                            if addr:
                                addresses.add(addr.lower())
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass

        return addresses

    def _subject_matches(self, item_subject, expected_subject):
        a = " ".join(str(item_subject or "").split()).lower()
        b = " ".join(str(expected_subject or "").split()).lower()
        if not b:
            return True
        return a == b or b in a or a in b

    def find_recent_sent(self, email, subject, max_check=80):
        """
        Look for a matching message in Sent Items by scanning recent items only.
        Avoids Sent Items.Count. Resolves Exchange SMTP addresses, not only To text.
        """
        email_l = (email or "").strip().lower()
        if not email_l:
            return False

        try:
            items = self.sent.Items
            try:
                items.Sort("[SentOn]", True)
            except Exception:
                try:
                    items.Sort("[LastModificationTime]", True)
                except Exception:
                    pass

            item = items.GetFirst()
            checked = 0
            while item is not None and checked < max_check:
                try:
                    addrs = self._recipient_addresses(item)
                    joined = " ".join(addrs)
                    recipient_ok = email_l in joined or any(email_l == a for a in addrs)
                    if recipient_ok and self._subject_matches(getattr(item, "Subject", ""), subject):
                        return True
                    # Recipient match alone is enough for very recent items (first 15)
                    if recipient_ok and checked < 15:
                        return True
                except Exception:
                    pass
                checked += 1
                try:
                    item = items.GetNext()
                except Exception:
                    break
        except Exception:
            pass
        return False

    def confirm_send(self, email, subject, outbox_before, timeout_sec=60):
        """
        Confirm Outlook sent the message primarily via Sent Items lookup.
        Outbox is only used as supporting evidence.
        """
        time.sleep(1.0)
        self.trigger_send_receive()

        seen_in_outbox = False
        outbox_cleared_after_seen = False
        deadline = time.time() + timeout_sec

        while time.time() < deadline:
            try:
                # Primary source of truth: Sent Items
                if self.find_recent_sent(email, subject):
                    return True, EmailState.CONFIRMED, None

                outbox_now = self.get_outbox_entry_ids()
                new_ids = outbox_now - outbox_before
                if new_ids:
                    seen_in_outbox = True
                    self.trigger_send_receive()
                elif seen_in_outbox:
                    outbox_cleared_after_seen = True
            except Exception:
                pass
            time.sleep(0.6)

        # Final Sent Items check (most important)
        if self.find_recent_sent(email, subject):
            return True, EmailState.CONFIRMED, None

        outbox_now = self.get_outbox_entry_ids()
        if outbox_now - outbox_before:
            return False, EmailState.UNKNOWN, "Email is still in Outlook Outbox — check internet/Outlook connection"

        if outbox_cleared_after_seen:
            # Left Outbox; Sent Items can lag on Exchange — treat as sent
            return True, EmailState.CONFIRMED, None

        return (
            False,
            EmailState.FAILED,
            "Could not verify in Sent Items. If the message is already there, "
            "recipient matching may have failed — check the Sent folder manually.",
        )

# =====================================================
# HELPER FUNCTION TO GET SURNAME
# =====================================================
def get_surname(fullname):
    """
    Returns the surname (part before comma) from a full name.
    If no comma is found, returns the full name.
    """
    fullname = str(fullname).strip()
    if "," in fullname:
        return fullname.split(",")[0].strip()
    return fullname

# =====================================================
# EMAIL VALIDATION
# =====================================================
def validate_email(email):
    """
    Validate email address format using regex
    Returns (is_valid: bool, error_message: str)
    """
    if not email or not str(email).strip():
        return False, "Email address is empty"
    
    email = str(email).strip()
    
    # Basic email regex pattern
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    
    if not re.match(pattern, email):
        return False, f"Invalid email format: {email}"
    
    # Additional checks
    if email.count('@') != 1:
        return False, f"Email must contain exactly one @ symbol: {email}"
    
    local, domain = email.split('@')
    
    if len(local) == 0:
        return False, "Local part of email is empty"
    
    if len(domain) == 0:
        return False, "Domain part of email is empty"
    
    if domain.startswith('.') or domain.endswith('.'):
        return False, f"Domain cannot start or end with dot: {domain}"
    
    return True, ""

def parse_cc_addresses(cc_string):
    """Parse multiple CC addresses from semicolon or comma separated string"""
    if not cc_string or not str(cc_string).strip():
        return []
    
    cc_string = str(cc_string).strip()
    # Try semicolon separator first, then comma
    if ';' in cc_string:
        addresses = [addr.strip() for addr in cc_string.split(';')]
    elif ',' in cc_string:
        addresses = [addr.strip() for addr in cc_string.split(',')]
    else:
        addresses = [cc_string]
    
    # Filter out empty addresses
    return [addr for addr in addresses if addr]

def validate_emails_in_dataframe(df, email_column="Email"):
    """
    Validate all emails in a dataframe column
    Returns (valid_df, invalid_emails: list)
    """
    if df is None or email_column not in df.columns:
        return df, []
    
    invalid_emails = []
    valid_indices = []
    
    for idx, row in df.iterrows():
        email = str(row[email_column]).strip()
        is_valid, error = validate_email(email)
        
        if is_valid or email == "":  # Allow empty emails (will be filtered later)
            valid_indices.append(idx)
        else:
            invalid_emails.append({
                'row': idx + 2,  # +2 for Excel row number (header + 1-based)
                'email': email,
                'error': error
            })
    
    valid_df = df.iloc[valid_indices].copy() if valid_indices else df.iloc[0:0].copy()
    return valid_df, invalid_emails

# =====================================================
# EXCEL VALIDATION
# =====================================================
def validate_excel_data(df):
    """
    Comprehensive Excel data validation before sending
    Returns (is_valid: bool, validation_results: dict)
    """
    validation_results = {
        'total_rows': len(df),
        'valid_rows': 0,
        'invalid_emails': [],
        'missing_attachments': [],
        'empty_required_fields': [],
        'duplicate_recipients': [],
        'invalid_cc': [],
        'warnings': []
    }
    
    if df is None or len(df) == 0:
        validation_results['warnings'].append("Excel file is empty")
        return False, validation_results
    
    # Check required columns
    required_columns = ['Account', 'Full Name', 'Email', 'CC', 'Attachment Path']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        validation_results['warnings'].append(f"Missing columns: {', '.join(missing_columns)}")
        return False, validation_results
    
    # Track seen emails for duplicate detection
    seen_emails = {}
    
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number
        has_issues = False
        
        # Validate email
        email = str(row['Email']).strip()
        if email:
            is_valid, error = validate_email(email)
            if not is_valid:
                validation_results['invalid_emails'].append({
                    'row': row_num,
                    'email': email,
                    'error': error
                })
                has_issues = True
            
            # Check for duplicates
            if email.lower() in seen_emails:
                validation_results['duplicate_recipients'].append({
                    'row': row_num,
                    'email': email,
                    'duplicate_row': seen_emails[email.lower()]
                })
                has_issues = True
            else:
                seen_emails[email.lower()] = row_num
        else:
            validation_results['empty_required_fields'].append({
                'row': row_num,
                'field': 'Email'
            })
            has_issues = True
        
        # Validate CC addresses
        cc_value = str(row['CC']).strip()
        if cc_value:
            cc_addresses = parse_cc_addresses(cc_value)
            for cc_addr in cc_addresses:
                is_valid, error = validate_email(cc_addr)
                if not is_valid:
                    validation_results['invalid_cc'].append({
                        'row': row_num,
                        'cc': cc_addr,
                        'error': error
                    })
                    has_issues = True
        
        # Validate attachment path
        attachment = str(row['Attachment Path']).strip()
        if attachment and not os.path.exists(attachment):
            validation_results['missing_attachments'].append({
                'row': row_num,
                'path': attachment
            })
            has_issues = True
        
        # Check other required fields
        if not str(row['Account']).strip():
            validation_results['empty_required_fields'].append({
                'row': row_num,
                'field': 'Account'
            })
            has_issues = True
        
        if not str(row['Full Name']).strip():
            validation_results['empty_required_fields'].append({
                'row': row_num,
                'field': 'Full Name'
            })
            has_issues = True
        
        if not has_issues:
            validation_results['valid_rows'] += 1
    
    # Determine overall validity
    is_valid = (
        len(validation_results['invalid_emails']) == 0 and
        len(validation_results['missing_attachments']) == 0 and
        len(validation_results['empty_required_fields']) == 0 and
        len(validation_results['invalid_cc']) == 0
    )
    
    return is_valid, validation_results

# =====================================================
# OUTLOOK-SAFE HTML BUILDER
# =====================================================
def build_outlook_safe_html(editor_html: str, para_spacing_px: int = 12) -> str:
    """
    Take rich HTML from QTextEdit.toHtml() and wrap it in an Outlook/Word-safe
    HTML shell with CSS resets to avoid extra spacing and reflow.
    """
    html = editor_html or ""
    PARA_SPACE_PX = max(0, int(para_spacing_px))

    # Extract inner <body> when present
    body_match = re.search(r"<body[^>]*>([\s\S]*?)</body>", html, re.IGNORECASE)
    inner = body_match.group(1) if body_match else html
    # Normalize divs to paragraphs
    inner = re.sub(r"<div\b([^>]*)>", r"<p\1>", inner, flags=re.IGNORECASE)
    inner = re.sub(r"</div>", r"</p>", inner, flags=re.IGNORECASE)
    had_paragraphs = bool(re.search(r"<p\b", inner, flags=re.IGNORECASE))
    # Convert multiple <br> into spacer blocks (Outlook-safe)
    inner = re.sub(
        r"(?:<br\s*/?>\s*){2,}",
        rf'''<table role="presentation" border="0" cellspacing="0" cellpadding="0" width="100%"><tr><td style="padding:0 0 {PARA_SPACE_PX}px 0;"><span style="font-size:1px; line-height:1px;">&nbsp;</span></td></tr></table>''',
        inner,
        flags=re.IGNORECASE
    )

    # Outlook spacing via tables
    # 1) Blank paragraphs -> spacer table
    inner = re.sub(
        r"<p\b[^>]*>\s*</p>",
        rf'''<table role="presentation" border="0" cellspacing="0" cellpadding="0" width="100%"><tr><td height="{PARA_SPACE_PX}" style="font-size:0; line-height:0;">&nbsp;</td></tr></table>''',
        inner,
        flags=re.IGNORECASE
    )
    # 2) Normal paragraphs -> table with content row + spacer row
    def _wrap_para(m):
        content = m.group(1)
        return (
            f'<table role="presentation" border="0" cellspacing="0" cellpadding="0" width="100%">'
            f'<tr><td style="line-height:1.35; mso-line-height-rule:exactly; font-family: Segoe UI, Arial, sans-serif;">{content}</td></tr>'
            f'<tr><td height="{PARA_SPACE_PX}" style="font-size:0; line-height:0;">&nbsp;</td></tr></table>'
        )
    inner = re.sub(
        r"<p\b[^>]*>([\s\S]*?)</p>",
        _wrap_para,
        inner,
        flags=re.IGNORECASE
    )
    # Fallback: if no paragraphs were present and no tables inserted, add spacing after single <br>
    if ('role="presentation"' not in inner) and (not had_paragraphs):
        inner = re.sub(
            r"<br\s*/?>",
            rf'''<br/><table role="presentation" border="0" cellspacing="0" cellpadding="0" width="100%"><tr><td height="{PARA_SPACE_PX}" style="font-size:0; line-height:0;">&nbsp;</td></tr></table>''',
            inner,
            flags=re.IGNORECASE
        )

    # Build final skeleton with resets
    wrapped = f"""<!DOCTYPE html>
<html xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns="http://www.w3.org/1999/xhtml">
<head>
  <meta http-equiv="x-ua-compatible" content="IE=edge">
  <meta name="format-detection" content="telephone=no, date=no, address=no, email=no">
  <meta name="x-apple-disable-message-reformatting">
  <!--[if mso]>
  <xml>
   <o:OfficeDocumentSettings>
    <o:AllowPNG/>
    <o:PixelsPerInch>96</o:PixelsPerInch>
   </o:OfficeDocumentSettings>
  </xml>
  <style type="text/css">
    body, table, td, div, p, a {{ font-family: Segoe UI, Arial, sans-serif !important; }}
    p {{ margin:0 !important; }}
  </style>
  <![endif]-->
  <style>
    body, table, td, div, p, a {{ font-family: Segoe UI, Arial, sans-serif; -webkit-text-size-adjust: 100%; -ms-text-size-adjust: 100%; }}
    p {{ margin:0 !important; }} /* kept for safety but spacing handled by tables */
    .content {{ font-size: 11pt; line-height: 1.35; color:#2b2b2b; }}
    img {{ border:0; outline:0; text-decoration:none; -ms-interpolation-mode:bicubic; }}
  </style>
</head>
<body style="Margin:0; padding:0; background:#ffffff;">
  <table role="presentation" cellpadding="0" cellspacing="0" border="0" width="100%">
    <tr>
      <td align="left" style="padding:0;">
        <div class="content" style="font-family: Segoe UI, Arial, sans-serif; font-size:11pt; line-height:1.35; mso-line-height-rule:exactly;">
          {inner}
        </div>
      </td>
    </tr>
  </table>
</body>
</html>"""
    return wrapped

# =====================================================
# EMAIL WORKER THREAD WITH ENHANCED SAFETY
# =====================================================
class EmailWorker(QThread):
    progress_updated = Signal(int)
    log_updated = Signal(str)
    status_updated = Signal(int, str)
    finished_sending = Signal()
    validation_complete = Signal(bool, object)  # is_valid, validation_results
    campaign_created = Signal(str)  # campaign_id

    def __init__(self, dataframe, subject, body_template, para_spacing_px=12, max_retries=3, 
                 send_delay=0, importance=2, request_read_receipt=True, campaign_name=None, db_manager=None, is_test_send=False):
        super().__init__()
        self.df = dataframe
        self.subject = subject
        self.body_template = body_template
        self.para_spacing_px = int(para_spacing_px) if para_spacing_px is not None else 12
        self.max_retries = max_retries
        self.send_delay = send_delay
        self.importance = importance  # 0=Low, 1=Normal, 2=High
        self.request_read_receipt = request_read_receipt
        self.running = True
        self.paused = False
        
        # Database manager for persistence
        self.db_manager = db_manager
        self.is_test_send = is_test_send
        
        # Campaign management
        self.campaign_id = None
        self.campaign_name = campaign_name or f"Campaign {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Track recipient states and attempts
        self.recipient_states = {}
        self.recipient_attempts = {}
        self.recipient_db_ids = {}  # Map row index to database ID
        
        # Initialize states
        for idx in range(len(dataframe)):
            self.recipient_states[idx] = EmailState.PENDING
            self.recipient_attempts[idx] = 0

    def stop(self):
        self.running = False

    def pause(self):
        self.paused = True

    def resume(self):
        self.paused = False

    def run(self):
        import pythoncom
        
        try:
            pythoncom.CoInitialize()
            
            try:
                # Phase 0: Create campaign in database (skip for test sends)
                if self.db_manager and not self.is_test_send:
                    # Generate unique campaign ID
                    date_str = datetime.now().strftime('%Y%m%d')
                    # Simple ID generation - in production you'd want something more robust
                    self.campaign_id = f"ESM-{date_str}-{int(time.time())}"
                    
                    # Create campaign
                    self.db_manager.create_campaign(self.campaign_id, self.campaign_name, len(self.df))
                    self.campaign_created.emit(self.campaign_id)
                    self.log_updated.emit(f"📋 Campaign created: {self.campaign_id}")
                    
                    # Add recipients to database
                    for index, row in self.df.iterrows():
                        recipient_id = self.db_manager.add_recipient(
                            self.campaign_id,
                            str(row.get("Account", "")),
                            str(row.get("Full Name", "")),
                            str(row.get("Email", "")),
                            str(row.get("CC", "")),
                            str(row.get("Attachment Path", "")),
                            index
                        )
                        if recipient_id:
                            self.recipient_db_ids[index] = recipient_id
                
                # Phase 1: Validation
                self.log_updated.emit("🔍 Starting Excel data validation...")
                is_valid, validation_results = validate_excel_data(self.df)
                self.validation_complete.emit(is_valid, validation_results)
                
                if not is_valid and not self.is_test_send:
                    self.log_updated.emit("❌ Validation failed. Please fix the errors before sending.")
                    if self.db_manager:
                        self.db_manager.update_campaign_status(self.campaign_id, 'validation_failed')
                    self.finished_sending.emit()
                    return
                
                self.log_updated.emit(f"✅ Validation complete. {validation_results['valid_rows']} valid rows ready to send.")
                
                # Phase 2: Connect to Outlook
                outlook_client = OutlookClient()
                success, message = outlook_client.connect()
                
                if not success:
                    self.log_updated.emit(f"❌ {message}")
                    self.log_updated.emit("Outlook is unavailable. Please make sure Microsoft Outlook is installed and configured, then try again.")
                    if self.db_manager:
                        self.db_manager.update_campaign_status(self.campaign_id, 'outlook_unavailable')
                    self.finished_sending.emit()
                    return
                
                self.log_updated.emit("✅ Connected to Outlook successfully.")
                
                # Phase 3: Send emails
                total = len(self.df)
                processed_count = 0
                
                for index, row in self.df.iterrows():
                    if not self.running:
                        self.log_updated.emit("⛔ Sending stopped by user.")
                        # Mark remaining as cancelled
                        for idx in range(index, len(self.df)):
                            if self.recipient_states[idx] == EmailState.PENDING:
                                self.recipient_states[idx] = EmailState.CANCELLED
                                self.status_updated.emit(idx, EmailState.CANCELLED)
                        break
                    
                    # Wait if paused
                    while self.paused and self.running:
                        time.sleep(0.1)
                    
                    if not self.running:
                        break
                    
                    # Skip if already processed
                    if EmailState.is_terminal(self.recipient_states[index]):
                        processed_count += 1
                        percent = int((processed_count / total) * 100)
                        self.progress_updated.emit(percent)
                        continue
                    
                    # Process this recipient
                    self.recipient_states[index] = EmailState.SENDING
                    self.status_updated.emit(index, EmailState.SENDING)
                    
                    result = self._send_single_email(outlook_client, row, index)
                    
                    if result['success']:
                        self.recipient_states[index] = EmailState.CONFIRMED
                        self.status_updated.emit(index, EmailState.CONFIRMED)
                        self.log_updated.emit(f"✅ Sent to {result['email']}")
                        
                        # Update database
                        if self.db_manager and index in self.recipient_db_ids:
                            recipient_id = self.recipient_db_ids[index]
                            self.db_manager.update_recipient_status(recipient_id, EmailState.CONFIRMED)
                            self.db_manager.log_send_attempt(recipient_id, self.campaign_id, 
                                                           self.recipient_attempts[index] + 1, 
                                                           EmailState.CONFIRMED)
                            
                    elif result['state'] == EmailState.UNKNOWN:
                        self.recipient_states[index] = EmailState.UNKNOWN
                        self.status_updated.emit(index, EmailState.UNKNOWN)
                        self.log_updated.emit(f"⚠️ Uncertain state for {result['email']}: {result['error']}")
                        
                        # Update database
                        if self.db_manager and index in self.recipient_db_ids:
                            recipient_id = self.recipient_db_ids[index]
                            self.db_manager.update_recipient_status(recipient_id, EmailState.UNKNOWN, result['error'])
                            self.db_manager.log_send_attempt(recipient_id, self.campaign_id,
                                                           self.recipient_attempts[index] + 1,
                                                           EmailState.UNKNOWN, result['error'])
                            
                    else:
                        self.recipient_states[index] = EmailState.FAILED
                        self.status_updated.emit(index, EmailState.FAILED)
                        self.log_updated.emit(f"❌ Failed to send to {result['email']}: {result['error']}")
                        
                        # Update database
                        if self.db_manager and index in self.recipient_db_ids:
                            recipient_id = self.recipient_db_ids[index]
                            self.db_manager.update_recipient_status(recipient_id, EmailState.FAILED, result['error'])
                            self.db_manager.log_send_attempt(recipient_id, self.campaign_id,
                                                           self.recipient_attempts[index] + 1,
                                                           EmailState.FAILED, result['error'])
                    
                    # Give Outlook time to process between messages
                    if index < len(self.df) - 1:
                        time.sleep(max(float(self.send_delay), 0.5))
                    
                    processed_count += 1
                    percent = int((processed_count / total) * 100)
                    self.progress_updated.emit(percent)
                
                # Phase 4: Retry failed emails (only true failures, not unknown)
                failed_indices = [idx for idx, state in self.recipient_states.items() 
                                if state == EmailState.FAILED and self.recipient_attempts[idx] < self.max_retries]
                
                if failed_indices and self.max_retries > 0:
                    self.log_updated.emit(f"🔄 Retrying {len(failed_indices)} failed emails...")
                    
                    for retry_attempt in range(self.max_retries):
                        if not self.running or not failed_indices:
                            break
                        
                        self.log_updated.emit(f"🔄 Retry attempt {retry_attempt + 1}/{self.max_retries}")
                        still_failed = []
                        
                        for index in failed_indices:
                            if not self.running:
                                break
                            
                            while self.paused and self.running:
                                time.sleep(0.1)
                            
                            if not self.running:
                                break
                            
                            row = self.df.iloc[index]
                            self.recipient_states[index] = EmailState.SENDING
                            self.status_updated.emit(index, EmailState.SENDING)
                            self.recipient_attempts[index] += 1
                            
                            result = self._send_single_email(outlook_client, row, index)
                            
                            if result['success']:
                                self.recipient_states[index] = EmailState.CONFIRMED
                                self.status_updated.emit(index, EmailState.CONFIRMED)
                                self.log_updated.emit(f"✅ Sent to {result['email']}")
                            elif result['state'] == EmailState.UNKNOWN:
                                self.recipient_states[index] = EmailState.UNKNOWN
                                self.status_updated.emit(index, EmailState.UNKNOWN)
                                still_failed.append(index)
                            else:
                                self.recipient_states[index] = EmailState.FAILED
                                self.status_updated.emit(index, EmailState.FAILED)
                                still_failed.append(index)
                            
                            time.sleep(max(float(self.send_delay), 0.5))
                        
                        failed_indices = still_failed
                        
                        if failed_indices:
                            time.sleep(2)  # Wait before next retry attempt
                
                # Summary
                confirmed = sum(1 for state in self.recipient_states.values() if state == EmailState.CONFIRMED)
                failed = sum(1 for state in self.recipient_states.values() if state == EmailState.FAILED)
                unknown = sum(1 for state in self.recipient_states.values() if state == EmailState.UNKNOWN)
                cancelled = sum(1 for state in self.recipient_states.values() if state == EmailState.CANCELLED)
                
                self.log_updated.emit(f"📊 Sending complete: {confirmed} sent, {failed} failed, {unknown} unknown, {cancelled} cancelled")
                
                if unknown > 0:
                    self.log_updated.emit("⚠️ Some emails have unknown status. These were not automatically retried to prevent duplicate sends.")
                
                # Update campaign in database
                if self.db_manager and not self.is_test_send:
                    self.db_manager.update_campaign_counts(self.campaign_id)
                    if cancelled > 0:
                        self.db_manager.update_campaign_status(self.campaign_id, 'cancelled')
                    else:
                        self.db_manager.update_campaign_status(self.campaign_id, 'completed')
                
                self.finished_sending.emit()
                
            finally:
                outlook_client.disconnect()
                pythoncom.CoUninitialize()
                
        except Exception as e:
            self.log_updated.emit(f"FATAL ERROR in worker: {str(e)}")
            self.finished_sending.emit()

    def _send_single_email(self, outlook_client, row, index):
        """Send a single email with enhanced error handling and state tracking"""
        try:
            email = str(row["Email"]).strip()
            cc_value = str(row["CC"]).strip()
            attachment = str(row["Attachment Path"]).strip()
            
            if not email:
                return {'success': False, 'state': EmailState.FAILED, 'email': email, 'error': 'No email address'}
            
            # Check attachment (should have been validated, but double-check)
            if attachment and not os.path.exists(attachment):
                return {'success': False, 'state': EmailState.FAILED, 'email': email, 'error': f'Attachment not found: {attachment}'}
            
            # Prepare email content
            account = str(row["Account"]).strip()
            full_name = str(row["Full Name"]).strip()
            surname = get_surname(full_name)
            
            # Replace placeholders
            subject = self.subject.replace("{{account}}", account).replace("{{Account}}", account).replace("{{fullname}}", full_name)
            body_raw = self.body_template.replace("{{fullname}}", surname)
            body = build_outlook_safe_html(body_raw, self.para_spacing_px)
            
            # Create email
            mail = outlook_client.create_email()
            mail.To = email
            
            # Handle CC addresses
            if cc_value:
                cc_addresses = parse_cc_addresses(cc_value)
                valid_cc = []
                for cc_addr in cc_addresses:
                    is_valid, _ = validate_email(cc_addr)
                    if is_valid:
                        valid_cc.append(cc_addr)
                if valid_cc:
                    mail.CC = ";".join(valid_cc)
            
            mail.Subject = subject
            
            try:
                mail.BodyFormat = 2  # olFormatHTML
            except Exception:
                pass
            
            mail.HTMLBody = body
            mail.Importance = self.importance  # 0=Low, 1=Normal, 2=High
            mail.ReadReceiptRequested = self.request_read_receipt
            
            # Add attachment if path provided
            if attachment:
                mail.Attachments.Add(attachment)

            try:
                mail.Save()
            except Exception:
                pass

            outbox_before = outlook_client.get_outbox_entry_ids()
            mail.Send()

            ok, state, error = outlook_client.confirm_send(email, subject, outbox_before)
            return {'success': ok, 'state': state, 'email': email, 'error': error}
            
        except Exception as e:
            error_msg = str(e)
            # Check if this might be a transient error
            if "unavailable" in error_msg.lower() or "not ready" in error_msg.lower():
                return {'success': False, 'state': EmailState.UNKNOWN, 'email': email, 'error': error_msg}
            else:
                return {'success': False, 'state': EmailState.FAILED, 'email': email, 'error': error_msg}


# =====================================================
# MAIN APPLICATION
# =====================================================
class EmailApp(QWidget):
    def __init__(self):
        super().__init__()

        self.db_manager = DatabaseManager()
        self.settings = SettingsManager()
        self.setup_logging()
        self.check_interrupted_campaigns()

        self._theme = self.settings.get("theme", "light")
        self.setWindowTitle("Eru Email Sender Pro")
        self.setMinimumSize(1280, 720)
        self.apply_theme()

        window_icon = self.create_app_icon()
        self.setWindowIcon(window_icon)

        app_icon = QIcon()
        icon_sizes = [16, 32, 48, 64, 128, 256]
        icon_base_path = "EMAIL.ico"
        if getattr(sys, 'frozen', False):
            if hasattr(sys, '_MEIPASS'):
                icon_base_path = os.path.join(sys._MEIPASS, "EMAIL.ico")
            else:
                app_dir = os.path.dirname(sys.executable)
                icon_base_path = os.path.join(app_dir, "EMAIL.ico")
        for size in icon_sizes:
            app_icon.addFile(icon_base_path, QSize(size, size))
        QApplication.setWindowIcon(app_icon)

        self.showMaximized()

        # Application shell
        shell_layout = QHBoxLayout(self)
        shell_layout.setContentsMargins(0, 0, 0, 0)
        shell_layout.setSpacing(0)

        self.sidebar = AppSidebar()
        self.sidebar.navigate.connect(self.navigate_to)
        self.sidebar.settings_btn.clicked.connect(lambda: self.navigate_to("settings"))
        shell_layout.addWidget(self.sidebar)

        self.main_content = QWidget()
        self.main_content.setObjectName("mainContent")
        main_layout = QVBoxLayout(self.main_content)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.page_stack = QStackedWidget()
        main_layout.addWidget(self.page_stack)
        shell_layout.addWidget(self.main_content, 1)

        # Build all pages
        self._build_dashboard_page()
        self._build_compose_page()
        self._build_recipients_page()
        self._build_templates_page()
        self._build_history_page()
        self._build_logs_page()
        self._build_settings_page()

        self.toast = ToastManager(self)

        # Wire core sending controls
        self.export_button.clicked.connect(self.export_template)
        self.import_button.clicked.connect(self.load_excel)
        self.start_button.clicked.connect(self.start_sending)
        self.stop_button.clicked.connect(self.stop_sending)
        self.pause_button.clicked.connect(self.pause_sending)
        self.resume_button.clicked.connect(self.resume_sending)
        self.retry_failed_button.clicked.connect(self.retry_failed_emails)
        self.test_send_button.clicked.connect(self.test_send)
        self.preview_button.clicked.connect(self.preview_email)
        self.validate_button.clicked.connect(self.validate_recipients)

        self.df = None
        self.worker = None
        self.current_campaign_id = None

        self.stats_timer = QTimer()
        self.stats_timer.timeout.connect(self.update_statistics)

        self.outlook_timer = QTimer()
        self.outlook_timer.timeout.connect(self.check_outlook_status)
        self.outlook_timer.start(15000)
        QTimer.singleShot(500, self.check_outlook_status)

        self.load_templates()
        self.template_combo.blockSignals(False)
        self.setup_keyboard_shortcuts()
        self.update_ui_state()
        self.navigate_to("dashboard")
        self.refresh_dashboard()

    def apply_theme(self):
        self.setStyleSheet(build_stylesheet(self._theme))

    def navigate_to(self, page_id: str):
        pages = {
            "dashboard": 0, "compose": 1, "recipients": 2,
            "templates": 3, "history": 4, "logs": 5, "settings": 6,
        }
        if page_id in pages:
            self.page_stack.setCurrentIndex(pages[page_id])
            self.sidebar.set_active(page_id)
            if page_id == "dashboard":
                self.refresh_dashboard()
            elif page_id == "history":
                self.refresh_history_page()
            elif page_id == "templates":
                self.refresh_templates_page()

    def check_outlook_status(self):
        try:
            client = OutlookClient()
            ok, msg = client.connect()
            if ok:
                self.sidebar.set_outlook_status(True, "Connected")
                client.disconnect()
            else:
                self.sidebar.set_outlook_status(False, "Not available")
        except Exception:
            self.sidebar.set_outlook_status(False, "Not available")

    def show_toast(self, message: str, toast_type: str = "info"):
        self.toast.show(message, toast_type)

    def _page_wrapper(self, header: PageHeader, body: QWidget) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(SPACING["xl"], SPACING["lg"], SPACING["xl"], SPACING["lg"])
        layout.setSpacing(SPACING["md"])
        layout.addWidget(header)
        layout.addWidget(body, 1)
        return page

    def _build_dashboard_page(self):
        header = PageHeader("Dashboard", "Overview of your email campaigns and activity")

        scroll = ContentScrollArea()
        body = scroll.content
        body_layout = scroll.layout

        # Stat cards row
        stats_row = QHBoxLayout()
        stats_row.setSpacing(SPACING["md"])
        self.dash_stat_sent = StatCard("Emails Sent", "0", LIGHT.success if self._theme == "light" else DARK.success)
        self.dash_stat_pending = StatCard("Pending", "0", LIGHT.accent if self._theme == "light" else DARK.accent)
        self.dash_stat_failed = StatCard("Failed", "0", LIGHT.error if self._theme == "light" else DARK.error)
        self.dash_stat_campaigns = StatCard("Campaigns", "0", "#8764B8")
        for card in [self.dash_stat_sent, self.dash_stat_pending, self.dash_stat_failed, self.dash_stat_campaigns]:
            stats_row.addWidget(card)
        body_layout.addLayout(stats_row)

        # Current session progress
        progress_section = make_section("Current Session", "Progress for the active recipient list")
        body_layout.addLayout(progress_section)
        progress_frame = QFrame()
        progress_frame.setObjectName("surfacePanel")
        pf_layout = QVBoxLayout(progress_frame)
        pf_layout.setContentsMargins(SPACING["lg"], SPACING["md"], SPACING["lg"], SPACING["md"])
        self.dashboard_progress = QProgressBar()
        self.dashboard_progress.setObjectName("progressBarLarge")
        self.dashboard_progress.setTextVisible(True)
        self.dashboard_progress.setFormat("Ready — load recipients to begin")
        self.dashboard_progress.setValue(0)
        pf_layout.addWidget(self.dashboard_progress)
        body_layout.addWidget(progress_frame)

        # Recent activity
        activity_section = make_section("Recent Campaigns", "Latest campaigns from your send history")
        body_layout.addLayout(activity_section)
        self.dashboard_campaigns_list = QFrame()
        self.dashboard_campaigns_list.setObjectName("surfacePanel")
        dcl_layout = QVBoxLayout(self.dashboard_campaigns_list)
        dcl_layout.setContentsMargins(SPACING["lg"], SPACING["md"], SPACING["lg"], SPACING["md"])
        self.dashboard_activity_label = QLabel("No campaigns yet. Start by importing recipients and sending emails.")
        self.dashboard_activity_label.setObjectName("sectionDesc")
        self.dashboard_activity_label.setWordWrap(True)
        dcl_layout.addWidget(self.dashboard_activity_label)
        body_layout.addWidget(self.dashboard_campaigns_list)

        # Quick actions
        actions_section = make_section("Quick Actions")
        body_layout.addLayout(actions_section)
        actions_row = QHBoxLayout()
        actions_row.setSpacing(SPACING["sm"])
        qa_import = QPushButton("Import Recipients")
        qa_import.setObjectName("primaryButton")
        qa_import.clicked.connect(lambda: (self.navigate_to("recipients"), self.load_excel()))
        qa_compose = QPushButton("Compose Email")
        qa_compose.setObjectName("secondaryButton")
        qa_compose.clicked.connect(lambda: self.navigate_to("compose"))
        qa_history = QPushButton("View History")
        qa_history.setObjectName("secondaryButton")
        qa_history.clicked.connect(lambda: self.navigate_to("history"))
        actions_row.addWidget(qa_import)
        actions_row.addWidget(qa_compose)
        actions_row.addWidget(qa_history)
        actions_row.addStretch()
        body_layout.addLayout(actions_row)
        body_layout.addStretch()

        page = self._page_wrapper(header, scroll)
        self.page_stack.addWidget(page)

    def refresh_dashboard(self):
        try:
            # Repair stale campaign counts (older rows used mismatched status casing)
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM campaigns")
            campaign_ids = [row[0] for row in cursor.fetchall()]
            conn.close()
            for campaign_id in campaign_ids:
                self.db_manager.update_campaign_counts(campaign_id)

            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM campaigns")
            campaign_count = cursor.fetchone()[0]
            cursor.execute("SELECT COALESCE(SUM(confirmed_count), 0) FROM campaigns")
            total_sent = cursor.fetchone()[0]
            cursor.execute("SELECT COALESCE(SUM(failed_count), 0) FROM campaigns")
            total_failed = cursor.fetchone()[0]
            cursor.execute('''
                SELECT name, total_recipients, confirmed_count, failed_count, status, started_at
                FROM campaigns ORDER BY started_at DESC LIMIT 5
            ''')
            recent = cursor.fetchall()
            conn.close()

            self.dash_stat_campaigns.set_value(str(campaign_count))
            self.dash_stat_sent.set_value(str(total_sent))
            self.dash_stat_failed.set_value(str(total_failed))

            if self.df is not None and len(self.df) > 0:
                pending = sum(1 for idx in range(len(self.df)) if str(self.df.iloc[idx, 5]).lower() == "pending")
                self.dash_stat_pending.set_value(str(pending))
            else:
                conn2 = self.db_manager.get_connection()
                cursor2 = conn2.cursor()
                cursor2.execute('''
                    SELECT COALESCE(SUM(total_recipients - confirmed_count - failed_count - unknown_count - cancelled_count), 0)
                    FROM campaigns WHERE status = 'in_progress'
                ''')
                pending = cursor2.fetchone()[0]
                conn2.close()
                self.dash_stat_pending.set_value(str(pending))

            if recent:
                lines = []
                for name, total, confirmed, failed, status, started in recent:
                    pct = int((confirmed / total) * 100) if total else 0
                    lines.append(f"{name}  —  {confirmed}/{total} sent ({pct}%)  ·  {status}")
                self.dashboard_activity_label.setText("\n".join(lines))
            else:
                self.dashboard_activity_label.setText("No campaigns yet. Start by importing recipients and sending emails.")
        except Exception:
            pass

    def _build_compose_page(self):
        header = PageHeader("Compose Email", "Create and send personalized email campaigns")
        actions = QHBoxLayout()
        actions.setSpacing(SPACING["sm"])
        self.preview_button = QPushButton("Preview")
        self.preview_button.setObjectName("secondaryButton")
        self.test_send_button = QPushButton("Send Test")
        self.test_send_button.setObjectName("secondaryButton")
        self.start_button = QPushButton("Send")
        self.start_button.setObjectName("successButton")
        self.pause_button = QPushButton("Pause")
        self.pause_button.setObjectName("secondaryButton")
        self.pause_button.setEnabled(False)
        self.resume_button = QPushButton("Resume")
        self.resume_button.setObjectName("secondaryButton")
        self.resume_button.setEnabled(False)
        self.stop_button = QPushButton("Stop")
        self.stop_button.setObjectName("dangerButton")
        self.stop_button.setEnabled(False)
        for btn in [self.preview_button, self.test_send_button, self.start_button,
                    self.pause_button, self.resume_button, self.stop_button]:
            header.add_action(btn)

        composer_body = self.create_email_panel()
        page = self._page_wrapper(header, composer_body)
        self.page_stack.addWidget(page)

    def _build_recipients_page(self):
        header = PageHeader("Recipients", "Manage and validate your recipient list")
        self.export_button = QPushButton("Export Template")
        self.export_button.setObjectName("secondaryButton")
        self.import_button = QPushButton("Import Excel")
        self.import_button.setObjectName("primaryButton")
        self.load_button = self.import_button
        self.validate_button = QPushButton("Validate")
        self.validate_button.setObjectName("secondaryButton")
        self.retry_failed_button = QPushButton("Retry Failed")
        self.retry_failed_button.setObjectName("secondaryButton")
        self.retry_failed_button.setEnabled(False)
        for btn in [self.export_button, self.import_button, self.validate_button, self.retry_failed_button]:
            header.add_action(btn)

        recipients_body = QWidget()
        rl = QVBoxLayout(recipients_body)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(SPACING["md"])

        # Search bar
        search_row = QHBoxLayout()
        self.recipient_search = QLineEdit()
        self.recipient_search.setObjectName("searchInput")
        self.recipient_search.setPlaceholderText("Search recipients by name, email, or account...")
        self.recipient_search.textChanged.connect(self.filter_recipients_table)
        search_row.addWidget(self.recipient_search)
        self.recipient_counter = QLabel("0 recipients")
        self.recipient_counter.setObjectName("sectionDesc")
        search_row.addWidget(self.recipient_counter)
        rl.addLayout(search_row)

        # Stats mini row
        stats_row = QHBoxLayout()
        stats_row.setSpacing(SPACING["md"])
        self.stat_total, self.stat_total_value = self._make_inline_stat("Total", "0")
        self.stat_sent, self.stat_sent_value = self._make_inline_stat("Sent", "0")
        self.stat_failed, self.stat_failed_value = self._make_inline_stat("Failed", "0")
        self.stat_pending, self.stat_pending_value = self._make_inline_stat("Pending", "0")
        for w in [self.stat_total, self.stat_sent, self.stat_failed, self.stat_pending]:
            stats_row.addWidget(w)
        stats_row.addStretch()
        rl.addLayout(stats_row)

        self.table = QTableWidget()
        self.table.setObjectName("dataTable")
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        rl.addWidget(self.table, 1)

        page = self._page_wrapper(header, recipients_body)
        self.page_stack.addWidget(page)

    def _make_inline_stat(self, label, value):
        card = QFrame()
        card.setObjectName("statCard")
        layout = QHBoxLayout(card)
        layout.setContentsMargins(SPACING["md"], SPACING["sm"], SPACING["md"], SPACING["sm"])
        lbl = QLabel(label)
        lbl.setObjectName("statLabel")
        val = QLabel(value)
        val.setObjectName("statValue")
        val.setStyleSheet("font-size: 14pt;")
        layout.addWidget(lbl)
        layout.addWidget(val)
        return card, val

    def filter_recipients_table(self, text):
        if self.df is None:
            return
        text = text.lower()
        for i in range(self.table.rowCount()):
            match = False
            for j in range(min(5, self.table.columnCount())):
                item = self.table.item(i, j)
                if item and text in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(i, bool(text) and not match)

    def validate_recipients(self):
        if self.df is None or len(self.df) == 0:
            self.show_toast("No recipients loaded. Import an Excel file first.", "warning")
            return
        is_valid, results = validate_excel_data(self.df)
        if is_valid:
            self.show_toast(f"All {results['valid_rows']} recipients passed validation.", "success")
        else:
            issues = []
            if results['invalid_emails']:
                issues.append(f"{len(results['invalid_emails'])} invalid emails")
            if results['missing_attachments']:
                issues.append(f"{len(results['missing_attachments'])} missing attachments")
            if results['empty_required_fields']:
                issues.append(f"{len(results['empty_required_fields'])} empty fields")
            if results['invalid_cc']:
                issues.append(f"{len(results['invalid_cc'])} invalid CC addresses")
            if results['duplicate_recipients']:
                issues.append(f"{len(results['duplicate_recipients'])} duplicates")
            self.show_toast(f"Validation issues: {', '.join(issues)}", "warning")

    def _build_templates_page(self):
        header = PageHeader("Templates", "Manage reusable email templates")
        self.new_template_btn = QPushButton("New Template")
        self.new_template_btn.setObjectName("primaryButton")
        self.new_template_btn.clicked.connect(self.save_template)
        header.add_action(self.new_template_btn)

        body = QWidget()
        bl = QHBoxLayout(body)
        bl.setContentsMargins(0, 0, 0, 0)
        bl.setSpacing(SPACING["md"])

        # Template list panel
        list_panel = QFrame()
        list_panel.setObjectName("surfacePanel")
        list_panel.setFixedWidth(280)
        lp_layout = QVBoxLayout(list_panel)
        lp_layout.setContentsMargins(SPACING["md"], SPACING["md"], SPACING["md"], SPACING["md"])
        self.template_search = QLineEdit()
        self.template_search.setPlaceholderText("Search templates...")
        self.template_search.textChanged.connect(self.filter_template_list)
        lp_layout.addWidget(self.template_search)
        from PySide6.QtWidgets import QListWidget
        self.template_list_widget = QListWidget()
        self.template_list_widget.currentRowChanged.connect(self._on_template_list_select)
        lp_layout.addWidget(self.template_list_widget, 1)
        bl.addWidget(list_panel)

        # Template preview panel
        preview_panel = QFrame()
        preview_panel.setObjectName("surfacePanel")
        pp_layout = QVBoxLayout(preview_panel)
        pp_layout.setContentsMargins(SPACING["lg"], SPACING["md"], SPACING["lg"], SPACING["md"])
        self.template_preview_subject = QLabel("")
        self.template_preview_subject.setObjectName("sectionTitle")
        self.template_preview_subject.setWordWrap(True)
        self.template_preview_body = QTextEdit()
        self.template_preview_body.setReadOnly(True)
        self.template_preview_body.setMinimumHeight(300)
        pp_layout.addWidget(make_field_label("Subject"))
        pp_layout.addWidget(self.template_preview_subject)
        pp_layout.addWidget(make_field_label("Body Preview"))
        pp_layout.addWidget(self.template_preview_body, 1)

        tpl_actions = QHBoxLayout()
        self.use_template_btn = QPushButton("Use in Composer")
        self.use_template_btn.setObjectName("primaryButton")
        self.use_template_btn.clicked.connect(self._use_selected_template)
        self.delete_template_page_btn = QPushButton("Delete")
        self.delete_template_page_btn.setObjectName("dangerButton")
        self.delete_template_page_btn.clicked.connect(self._delete_template_from_page)
        tpl_actions.addWidget(self.use_template_btn)
        tpl_actions.addWidget(self.delete_template_page_btn)
        tpl_actions.addStretch()
        pp_layout.addLayout(tpl_actions)
        bl.addWidget(preview_panel, 1)

        page = self._page_wrapper(header, body)
        self.page_stack.addWidget(page)

    def refresh_templates_page(self):
        self.template_list_widget.clear()
        self.template_list_widget.addItem("Default HR Notice")
        templates = self.settings.get("email_templates", {})
        for name in sorted(templates.keys()):
            self.template_list_widget.addItem(name)
        if self.template_list_widget.count() > 0:
            self.template_list_widget.setCurrentRow(0)

    def filter_template_list(self, text):
        text = text.lower()
        for i in range(self.template_list_widget.count()):
            item = self.template_list_widget.item(i)
            item.setHidden(bool(text) and text not in item.text().lower())

    def _on_template_list_select(self, row):
        if row < 0:
            return
        name = self.template_list_widget.item(row).text()
        if name == "Default HR Notice":
            self.template_preview_subject.setText("NOTICE TO SUBMIT LACKING EMPLOYMENT REQUIREMENTS - {{fullname}} - {{account}}")
            self.template_preview_body.setPlainText("Default HR notice template with {{fullname}} and {{account}} placeholders.")
        else:
            templates = self.settings.get("email_templates", {})
            if name in templates:
                tpl = templates[name]
                self.template_preview_subject.setText(tpl.get("subject", ""))
                body_html = tpl.get("body", "")
                self.template_preview_body.setHtml(body_html)

    def _use_selected_template(self):
        row = self.template_list_widget.currentRow()
        if row < 0:
            return
        name = self.template_list_widget.item(row).text()
        for i in range(self.template_combo.count()):
            if self.template_combo.itemText(i) == name or (name == "Default HR Notice" and i == 0):
                self.template_combo.setCurrentIndex(i)
                break
        self.navigate_to("compose")
        self.show_toast(f"Template '{name}' loaded in composer.", "success")

    def _build_history_page(self):
        header = PageHeader("History", "Campaign send history and details")
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setObjectName("secondaryButton")
        refresh_btn.clicked.connect(self.refresh_history_page)
        header.add_action(refresh_btn)

        body = QWidget()
        bl = QVBoxLayout(body)
        bl.setContentsMargins(0, 0, 0, 0)
        bl.setSpacing(SPACING["md"])

        self.history_search = QLineEdit()
        self.history_search.setPlaceholderText("Search campaigns...")
        self.history_search.textChanged.connect(self.filter_history_table)
        bl.addWidget(self.history_search)

        self.history_table = QTableWidget()
        self.history_table.setColumnCount(7)
        self.history_table.setHorizontalHeaderLabels(
            ["Campaign", "Status", "Total", "Sent", "Failed", "Started", "Completed"]
        )
        self.history_table.setAlternatingRowColors(True)
        self.history_table.setShowGrid(False)
        self.history_table.verticalHeader().setVisible(False)
        self.history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.history_table.horizontalHeader().setStretchLastSection(True)
        self.history_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.history_table.doubleClicked.connect(self._history_row_details)
        bl.addWidget(self.history_table, 1)

        hist_actions = QHBoxLayout()
        self.history_view_btn = QPushButton("View Details")
        self.history_view_btn.setObjectName("secondaryButton")
        self.history_view_btn.clicked.connect(self._history_view_selected)
        self.history_delete_btn = QPushButton("Delete Campaign")
        self.history_delete_btn.setObjectName("dangerButton")
        self.history_delete_btn.clicked.connect(self._history_delete_selected)
        self.view_history_button = self.history_view_btn
        hist_actions.addWidget(self.history_view_btn)
        hist_actions.addWidget(self.history_delete_btn)
        hist_actions.addStretch()
        bl.addLayout(hist_actions)

        page = self._page_wrapper(header, body)
        self.page_stack.addWidget(page)

    def refresh_history_page(self):
        try:
            conn = self.db_manager.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, name, total_recipients, confirmed_count, failed_count,
                       status, started_at, completed_at
                FROM campaigns ORDER BY started_at DESC LIMIT 100
            ''')
            campaigns = cursor.fetchall()
            conn.close()

            self.history_table.setRowCount(len(campaigns))
            for i, row in enumerate(campaigns):
                cid, name, total, confirmed, failed, status, started, completed = row
                display = f"{name} ({cid})"
                for j, val in enumerate([display, status, str(total), str(confirmed),
                                         str(failed), started or "", completed or ""]):
                    item = QTableWidgetItem(str(val))
                    item.setData(Qt.UserRole, cid if j == 0 else None)
                    if j == 0:
                        item.setData(Qt.UserRole, cid)
                    self.history_table.setItem(i, j, item)
        except Exception as e:
            self.show_toast(f"Failed to load history: {e}", "error")

    def filter_history_table(self, text):
        text = text.lower()
        for i in range(self.history_table.rowCount()):
            match = False
            for j in range(self.history_table.columnCount()):
                item = self.history_table.item(i, j)
                if item and text in item.text().lower():
                    match = True
                    break
            self.history_table.setRowHidden(i, bool(text) and not match)

    def _history_view_selected(self):
        row = self.history_table.currentRow()
        if row < 0:
            self.show_toast("Select a campaign to view details.", "warning")
            return
        cid = self.history_table.item(row, 0).data(Qt.UserRole)
        self._show_campaign_details(cid)

    def _history_row_details(self, index):
        cid = self.history_table.item(index.row(), 0).data(Qt.UserRole)
        self._show_campaign_details(cid)

    def _history_delete_selected(self):
        row = self.history_table.currentRow()
        if row < 0:
            self.show_toast("Select a campaign to delete.", "warning")
            return
        cid = self.history_table.item(row, 0).data(Qt.UserRole)
        reply = QMessageBox.question(self, "Delete Campaign",
                                     f"Delete campaign {cid}? This cannot be undone.",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            try:
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                cursor.execute('DELETE FROM send_logs WHERE campaign_id = ?', (cid,))
                cursor.execute('DELETE FROM send_attempts WHERE campaign_id = ?', (cid,))
                cursor.execute('DELETE FROM recipients WHERE campaign_id = ?', (cid,))
                cursor.execute('DELETE FROM campaigns WHERE id = ?', (cid,))
                conn.commit()
                conn.close()
                self.refresh_history_page()
                self.refresh_dashboard()
                self.show_toast("Campaign deleted.", "success")
            except Exception as e:
                self.show_toast(f"Failed to delete: {e}", "error")

    def _show_campaign_details(self, campaign_id):
        try:
            recipients = self.db_manager.get_campaign_recipients(campaign_id)
            if not recipients:
                self.show_toast("No recipients found for this campaign.", "warning")
                return
            from PySide6.QtWidgets import QDialog, QVBoxLayout, QDialogButtonBox
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Campaign Details — {campaign_id}")
            dialog.setMinimumSize(800, 600)
            dialog.setStyleSheet(build_stylesheet(self._theme))
            layout = QVBoxLayout(dialog)
            details = QTextEdit()
            details.setReadOnly(True)
            text = f"Campaign ID: {campaign_id}\n\n"
            for rec in recipients:
                (rec_id, account, full_name, email, cc, attachment_path, status,
                 attempt_number, last_error, row_index) = rec
                text += f"{full_name} ({email}) — {status}"
                if last_error:
                    text += f" — {last_error}"
                text += "\n"
            details.setPlainText(text)
            layout.addWidget(details)
            buttons = QDialogButtonBox(QDialogButtonBox.Close)
            buttons.rejected.connect(dialog.close)
            layout.addWidget(buttons)
            dialog.exec()
        except Exception as e:
            self.show_toast(f"Failed to load details: {e}", "error")

    def _build_logs_page(self):
        header = PageHeader("Logs", "Activity log for troubleshooting")
        clear_btn = QPushButton("Clear")
        clear_btn.setObjectName("secondaryButton")
        clear_btn.clicked.connect(lambda: self.log_box.clear())
        header.add_action(clear_btn)

        body = QWidget()
        bl = QVBoxLayout(body)
        bl.setContentsMargins(0, 0, 0, 0)
        bl.setSpacing(SPACING["md"])

        filter_row = QHBoxLayout()
        self.log_filter = QComboBox()
        self.log_filter.addItems(["All", "Info", "Success", "Warning", "Error"])
        self.log_filter.currentIndexChanged.connect(self._filter_logs)
        filter_row.addWidget(make_field_label("Filter"))
        filter_row.addWidget(self.log_filter)
        self.log_search = QLineEdit()
        self.log_search.setPlaceholderText("Search logs...")
        filter_row.addWidget(self.log_search, 1)
        bl.addLayout(filter_row)

        self.log_box = QTextEdit()
        self.log_box.setObjectName("logBox")
        self.log_box.setReadOnly(True)
        bl.addWidget(self.log_box, 1)

        # Progress bar for sending (also shown on logs page)
        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("progressBarLarge")
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("Ready")
        bl.addWidget(self.progress_bar)

        page = self._page_wrapper(header, body)
        self.page_stack.addWidget(page)

    def _filter_logs(self):
        pass  # Visual filter placeholder — logs remain full text for troubleshooting

    def _build_settings_page(self):
        header = PageHeader("Settings", "Configure application behavior and preferences")

        scroll = ContentScrollArea()
        body = scroll.content
        bl = scroll.layout

        # General section
        bl.addLayout(make_section("General", "Sending behavior and retry configuration"))
        general_frame = QFrame()
        general_frame.setObjectName("surfacePanel")
        gf = QFormLayout(general_frame)
        gf.setContentsMargins(SPACING["lg"], SPACING["md"], SPACING["lg"], SPACING["md"])
        gf.setSpacing(SPACING["md"])
        self.settings_max_retries = QSpinBox()
        self.settings_max_retries.setRange(0, 10)
        self.settings_max_retries.setValue(self.settings.get("max_retries", 3))
        self.settings_send_delay = QSpinBox()
        self.settings_send_delay.setRange(0, 60)
        self.settings_send_delay.setValue(self.settings.get("send_delay", 0))
        self.settings_send_delay.setSuffix(" sec")
        self.settings_auto_save = QSpinBox()
        self.settings_auto_save.setRange(1, 60)
        self.settings_auto_save.setValue(self.settings.get("auto_save_interval", 5))
        self.settings_auto_save.setSuffix(" min")
        gf.addRow("Max Retries:", self.settings_max_retries)
        gf.addRow("Send Delay:", self.settings_send_delay)
        gf.addRow("Auto-save Interval:", self.settings_auto_save)
        bl.addWidget(general_frame)

        # Email section
        bl.addLayout(make_section("Outlook & Email", "Email delivery settings"))
        email_frame = QFrame()
        email_frame.setObjectName("surfacePanel")
        ef = QFormLayout(email_frame)
        ef.setContentsMargins(SPACING["lg"], SPACING["md"], SPACING["lg"], SPACING["md"])
        self.settings_importance = QComboBox()
        self.settings_importance.addItem("Low", 0)
        self.settings_importance.addItem("Normal", 1)
        self.settings_importance.addItem("High", 2)
        current = self.settings.get("importance", 2)
        for i in range(self.settings_importance.count()):
            if self.settings_importance.itemData(i) == current:
                self.settings_importance.setCurrentIndex(i)
                break
        self.settings_read_receipt = QCheckBox("Request read receipt")
        self.settings_read_receipt.setChecked(self.settings.get("request_read_receipt", True))
        self.settings_retry_failed = QCheckBox("Retry failed emails automatically")
        self.settings_retry_failed.setChecked(self.settings.get("retry_failed_emails", True))
        ef.addRow("Email Importance:", self.settings_importance)
        ef.addRow("", self.settings_read_receipt)
        ef.addRow("", self.settings_retry_failed)
        bl.addWidget(email_frame)

        # Appearance section
        bl.addLayout(make_section("Appearance", "Theme and display preferences"))
        ui_frame = QFrame()
        ui_frame.setObjectName("surfacePanel")
        uf = QFormLayout(ui_frame)
        uf.setContentsMargins(SPACING["lg"], SPACING["md"], SPACING["lg"], SPACING["md"])
        self.settings_theme = QComboBox()
        self.settings_theme.addItem("Light", "light")
        self.settings_theme.addItem("Dark", "dark")
        for i in range(self.settings_theme.count()):
            if self.settings_theme.itemData(i) == self._theme:
                self.settings_theme.setCurrentIndex(i)
                break
        self.settings_spacing = QComboBox()
        self.settings_spacing.addItem("Tight", 8)
        self.settings_spacing.addItem("Normal", 12)
        self.settings_spacing.addItem("Relaxed", 16)
        sp = self.settings.get("paragraph_spacing", 12)
        for i in range(self.settings_spacing.count()):
            if self.settings_spacing.itemData(i) == sp:
                self.settings_spacing.setCurrentIndex(i)
                break
        uf.addRow("Theme:", self.settings_theme)
        uf.addRow("Paragraph Spacing:", self.settings_spacing)
        bl.addWidget(ui_frame)

        save_row = QHBoxLayout()
        save_settings_btn = QPushButton("Save Settings")
        save_settings_btn.setObjectName("primaryButton")
        save_settings_btn.clicked.connect(self._save_settings_page)
        save_row.addWidget(save_settings_btn)
        save_row.addStretch()
        bl.addLayout(save_row)
        bl.addStretch()

        self.settings_button = save_settings_btn
        page = self._page_wrapper(header, scroll)
        self.page_stack.addWidget(page)

    def _save_settings_page(self):
        self.settings.set("max_retries", self.settings_max_retries.value())
        self.settings.set("send_delay", self.settings_send_delay.value())
        self.settings.set("auto_save_interval", self.settings_auto_save.value())
        self.settings.set("importance", self.settings_importance.currentData())
        self.settings.set("request_read_receipt", self.settings_read_receipt.isChecked())
        self.settings.set("retry_failed_emails", self.settings_retry_failed.isChecked())
        self.settings.set("paragraph_spacing", self.settings_spacing.currentData())

        new_theme = self.settings_theme.currentData()
        if new_theme != self._theme:
            self._theme = new_theme
            self.settings.set("theme", new_theme)
            self.apply_theme()

        try:
            px = int(self.settings_spacing.currentData())
            self.apply_editor_paragraph_spacing(px)
            for i in range(self.spacing_select.count()):
                if self.spacing_select.itemData(i) == px:
                    self.spacing_select.setCurrentIndex(i)
                    break
        except Exception:
            pass

        self.show_toast("Settings saved successfully.", "success")

    def show_settings_dialog(self):
        self.navigate_to("settings")
    
    def check_interrupted_campaigns(self):
        """Check for interrupted campaigns and show recovery dialog"""
        try:
            interrupted = self.db_manager.get_interrupted_campaigns()
            if interrupted:
                self.show_recovery_dialog(interrupted)
        except Exception as e:
            print(f"Error checking interrupted campaigns: {e}")
    
    def show_recovery_dialog(self, campaigns):
        """Show dialog for interrupted campaign recovery"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QTextBrowser, QHBoxLayout, QLabel, QPushButton, QListWidget, QListWidgetItem
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Interrupted Campaigns Found")
        dialog.setMinimumSize(600, 400)
        dialog.setStyleSheet(build_stylesheet(self._theme))
        
        layout = QVBoxLayout(dialog)
        
        # Information label
        info_label = QLabel("The following campaigns were interrupted and can be resumed:")
        layout.addWidget(info_label)
        
        # Campaign list
        campaign_list = QListWidget()
        for campaign in campaigns:
            (campaign_id, name, total, confirmed, failed, unknown, cancelled, started_at) = campaign
            pending = total - confirmed - failed - unknown - cancelled
            item_text = f"{name} ({campaign_id})\n"
            item_text += f"  Total: {total} | Sent: {confirmed} | Failed: {failed} | Unknown: {unknown} | Pending: {pending}"
            item = QListWidgetItem(item_text)
            item.setData(Qt.UserRole, campaign_id)
            campaign_list.addItem(item)
        
        layout.addWidget(campaign_list)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        resume_button = QPushButton("Resume Selected")
        resume_button.setObjectName("successButton")
        resume_button.clicked.connect(lambda: self.resume_campaign(dialog, campaign_list))

        review_button = QPushButton("Review Details")
        review_button.setObjectName("primaryButton")
        review_button.clicked.connect(lambda: self.review_campaign(dialog, campaign_list))

        cancel_button = QPushButton("Cancel")
        cancel_button.setObjectName("secondaryButton")
        cancel_button.clicked.connect(dialog.close)
        
        button_layout.addWidget(resume_button)
        button_layout.addWidget(review_button)
        button_layout.addStretch()
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        
        dialog.exec()
    
    def resume_campaign(self, dialog, campaign_list):
        """Resume selected campaign"""
        selected_items = campaign_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "No Selection", "Please select a campaign to resume.")
            return
        
        campaign_id = selected_items[0].data(Qt.UserRole)
        dialog.close()
        
        # Load campaign data
        try:
            recipients = self.db_manager.get_campaign_recipients(campaign_id)
            if recipients:
                # Convert to dataframe
                data = []
                for rec in recipients:
                    (rec_id, account, full_name, email, cc, attachment_path, status, 
                     attempt_number, last_error, row_index) = rec
                    data.append({
                        'Account': account,
                        'Full Name': full_name,
                        'Email': email,
                        'CC': cc,
                        'Attachment Path': attachment_path,
                        'Status': status
                    })
                
                self.df = pd.DataFrame(data)
                self.populate_table()
                self.update_ui_state()
                
                QMessageBox.information(self, "Campaign Loaded", 
                                      f"Campaign {campaign_id} has been loaded. "
                                      f"You can review the status and send pending emails.")
                self.log(f"📥 Resumed campaign: {campaign_id}")
            else:
                QMessageBox.warning(self, "Error", "No recipients found for this campaign.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load campaign: {str(e)}")
    
    def review_campaign(self, dialog, campaign_list):
        """Review selected campaign details"""
        selected_items = campaign_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "No Selection", "Please select a campaign to review.")
            return
        
        campaign_id = selected_items[0].data(Qt.UserRole)
        
        # Show detailed information
        try:
            recipients = self.db_manager.get_campaign_recipients(campaign_id)
            if recipients:
                # Create a simple details dialog
                from PySide6.QtWidgets import QDialog, QVBoxLayout, QTextEdit
                
                details_dialog = QDialog(self)
                details_dialog.setWindowTitle(f"Campaign Details: {campaign_id}")
                details_dialog.setMinimumSize(800, 600)
                details_dialog.setStyleSheet(build_stylesheet(self._theme))
                
                layout = QVBoxLayout(details_dialog)
                
                details_text = QTextEdit()
                details_text.setReadOnly(True)
                
                # Build details text
                details = f"Campaign ID: {campaign_id}\n\n"
                details += "RECIPIENTS:\n"
                details += "-" * 80 + "\n"
                
                for rec in recipients:
                    (rec_id, account, full_name, email, cc, attachment_path, status, 
                     attempt_number, last_error, row_index) = rec
                    details += f"Row {row_index + 2}: {full_name} ({email})\n"
                    details += f"  Status: {status} | Attempts: {attempt_number}\n"
                    if last_error:
                        details += f"  Last Error: {last_error}\n"
                    details += "\n"
                
                details_text.setPlainText(details)
                layout.addWidget(details_text)
                
                close_button = QPushButton("Close")
                close_button.clicked.connect(details_dialog.close)
                layout.addWidget(close_button)
                
                details_dialog.exec()
            else:
                QMessageBox.warning(self, "Error", "No recipients found for this campaign.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load campaign details: {str(e)}")

    # =================================================
    # UI COMPONENT CREATION METHODS
    # =================================================
    def create_app_icon(self):
        """Create app icon using the EMAIL.ico file"""
        # Handle both script and executable environments
        icon_path = "EMAIL.ico"
        
        # If running as executable, check PyInstaller's temporary directory first
        if getattr(sys, 'frozen', False):
            # First try PyInstaller's temporary directory where bundled files are extracted
            if hasattr(sys, '_MEIPASS'):
                icon_path = os.path.join(sys._MEIPASS, "EMAIL.ico")
            else:
                # Fallback to executable directory
                app_dir = os.path.dirname(sys.executable)
                icon_path = os.path.join(app_dir, "EMAIL.ico")
        
        if os.path.exists(icon_path):
            return QIcon(icon_path)
        else:
            # Fallback to a simple colored icon if EMAIL.ico is not found
            pixmap = QPixmap(32, 32)
            pixmap.fill(QColor("#4a90e2"))
            return QIcon(pixmap)
    
    def create_email_panel(self):
        """Create the professional email composer panel."""
        email_frame = QWidget()
        email_layout = QVBoxLayout(email_frame)
        email_layout.setContentsMargins(0, 0, 0, 0)
        email_layout.setSpacing(SPACING["md"])

        # Subject + template row
        meta_frame = QFrame()
        meta_frame.setObjectName("surfacePanel")
        meta_layout = QVBoxLayout(meta_frame)
        meta_layout.setContentsMargins(SPACING["lg"], SPACING["md"], SPACING["lg"], SPACING["md"])
        meta_layout.setSpacing(SPACING["md"])

        template_row = QHBoxLayout()
        template_row.addWidget(make_field_label("Template"))
        self.template_combo = QComboBox()
        self.template_combo.setObjectName("templateCombo")
        self.template_combo.addItem("Default HR Notice", "default")
        self.template_combo.blockSignals(True)
        self.template_combo.currentIndexChanged.connect(self.load_template)
        template_row.addWidget(self.template_combo, 1)
        self.save_template_btn = QPushButton("Save")
        self.save_template_btn.setObjectName("secondaryButton")
        self.save_template_btn.clicked.connect(self.save_template)
        self.delete_template_btn = QPushButton("Delete")
        self.delete_template_btn.setObjectName("dangerButton")
        self.delete_template_btn.clicked.connect(self.delete_template)
        template_row.addWidget(self.save_template_btn)
        template_row.addWidget(self.delete_template_btn)
        meta_layout.addLayout(template_row)

        meta_layout.addWidget(make_field_label("Subject"))
        self.subject_input = QLineEdit()
        self.subject_input.setObjectName("subjectInput")
        self.subject_input.setPlaceholderText("Enter email subject — use {{fullname}} and {{account}} placeholders")
        self.subject_input.setText("NOTICE TO SUBMIT LACKING EMPLOYMENT REQUIREMENTS - {{fullname}} - {{account}}")
        meta_layout.addWidget(self.subject_input)
        email_layout.addWidget(meta_frame)

        # Formatting toolbar
        toolbar_frame = QFrame()
        toolbar_frame.setObjectName("toolbarFrame")
        tb_layout = QHBoxLayout(toolbar_frame)
        tb_layout.setContentsMargins(SPACING["sm"], SPACING["sm"], SPACING["sm"], SPACING["sm"])
        tb_layout.setSpacing(SPACING["xs"])

        self.bold_btn = QPushButton("B")
        self.bold_btn.setObjectName("toolbarButton")
        self.bold_btn.setFixedSize(32, 32)
        self.bold_btn.clicked.connect(self.make_bold)
        self.italic_btn = QPushButton("I")
        self.italic_btn.setObjectName("toolbarButton")
        self.italic_btn.setFixedSize(32, 32)
        self.italic_btn.clicked.connect(self.make_italic)
        self.underline_btn = QPushButton("U")
        self.underline_btn.setObjectName("toolbarButton")
        self.underline_btn.setFixedSize(32, 32)
        self.underline_btn.clicked.connect(self.make_underline)

        tb_layout.addWidget(self.bold_btn)
        tb_layout.addWidget(self.italic_btn)
        tb_layout.addWidget(self.underline_btn)

        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setStyleSheet("color: #EDEBE9;")
        tb_layout.addWidget(sep)

        # Variable insert menu
        self.variable_btn = QToolButton()
        self.variable_btn.setText("Insert Variable")
        self.variable_btn.setObjectName("toolbarButton")
        self.variable_btn.setPopupMode(QToolButton.InstantPopup)
        var_menu = QMenu(self.variable_btn)
        for var in ["{{fullname}}", "{{account}}", "{{Account}}"]:
            action = var_menu.addAction(var)
            action.triggered.connect(lambda checked, v=var: self._insert_variable(v))
        self.variable_btn.setMenu(var_menu)
        tb_layout.addWidget(self.variable_btn)

        tb_layout.addStretch()

        spacing_label = QLabel("Spacing")
        spacing_label.setObjectName("fieldLabel")
        self.spacing_select = QComboBox()
        self.spacing_select.setObjectName("spacingSelect")
        self.spacing_select.addItem("Tight", 8)
        self.spacing_select.addItem("Normal", 12)
        self.spacing_select.addItem("Relaxed", 16)
        saved_spacing = self.settings.get("paragraph_spacing", 12)
        for i in range(self.spacing_select.count()):
            if self.spacing_select.itemData(i) == saved_spacing:
                self.spacing_select.setCurrentIndex(i)
                break
        self.spacing_select.currentIndexChanged.connect(self.on_spacing_changed)
        tb_layout.addWidget(spacing_label)
        tb_layout.addWidget(self.spacing_select)
        email_layout.addWidget(toolbar_frame)

        # Email body editor
        editor_frame = QFrame()
        editor_frame.setObjectName("editorFrame")
        editor_layout = QVBoxLayout(editor_frame)
        editor_layout.setContentsMargins(0, 0, 0, 0)
        self.email_editor = QTextEdit()
        self.email_editor.setObjectName("emailEditor")
        self.email_editor.setFont(QFont("Segoe UI", 11))
        self.email_editor.setMinimumHeight(400)
        self.email_editor.setAcceptRichText(True)
        self.email_editor.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.email_editor.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        default_body = """
<p>Dear {{fullname}},</p>

<p>This is to formally inform you that you still have outstanding mandatory employment requirements as of this date, despite prior reminders and your signed Affidavit of Undertaking upon commencement of employment.</p>

<p>As stated in your Affidavit of Undertaking, you committed to submit all required documents within prescribed period. <b>You are hereby given five (5) days from receipt of this email notice </b> to complete and submit pending requirements. Please see the attached <b>Notice of Incomplete Employment Requirements</b> for full details. Failure to comply within the given timeframe, may result in appropriate administrative action in accordance with Company policy.</p>

<p>Please submit the required documents through this same email thread. For any clarification, please coordinate with <b>HR-DMRC or your assigned account supervisor.</b></p>

<p>Thanks,<br>Jhudel S. Orola<br>HR Staff - Data Management & Records Control<br>Acabar Marketing International Inc.<br>(02) 8887-8170 Local 153</p>

<p><img class="x_CToWUd" height="77" width="250" src="https://ci3.googleusercontent.com/mail-sig/AIorK4x0oCXqeBBsjR9hQB3HLxhAJPc1msod_2dqrIiATYz-sDfATgJdOa_R6eWlr16--ykbMmeApG_G3we-" data-imagetype="External"></p>
"""
        self.email_editor.setHtml(default_body)
        try:
            self.apply_editor_paragraph_spacing(int(self.spacing_select.currentData()))
        except Exception:
            self.apply_editor_paragraph_spacing(12)

        editor_layout.addWidget(self.email_editor)
        email_layout.addWidget(editor_frame, 1)

        # Placeholder hint
        hint = QLabel("Supported placeholders: {{fullname}} (surname in body, full name in subject), {{account}}, {{Account}}")
        hint.setObjectName("sectionDesc")
        hint.setWordWrap(True)
        email_layout.addWidget(hint)

        return email_frame

    def _insert_variable(self, variable: str):
        cursor = self.email_editor.textCursor()
        cursor.insertText(variable)
        self.email_editor.setFocus()
    
    def update_ui_state(self):
        """Update UI state based on data availability"""
        has_data = self.df is not None and len(self.df) > 0
        self.start_button.setEnabled(has_data)
        self.stop_button.setEnabled(False)
        self.pause_button.setEnabled(False)
        self.resume_button.setEnabled(False)
        
        # Check if there are failed emails to retry
        has_failed = False
        if has_data:
            for idx in range(len(self.df)):
                status = str(self.df.iloc[idx, 5]).lower()
                if status == "failed":
                    has_failed = True
                    break
        self.retry_failed_button.setEnabled(has_failed)
        
        # Update statistics cards
        if has_data:
            self.update_statistics()
        else:
            self.stat_total_value.setText("0")
            self.stat_sent_value.setText("0")
            self.stat_failed_value.setText("0")
            self.stat_pending_value.setText("0")
            self.recipient_counter.setText("0 recipients")

    # =================================================
    # EXPORT EXCEL TEMPLATE
    # =================================================
    def export_template(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel Template", "Email_Template.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Email Sending Setup"
            headers = ["Account", "Full Name", "Email", "CC", "Attachment Path"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            ws.cell(row=2, column=1).value = "ACC-001"
            ws.cell(row=2, column=2).value = "Dela Cruz, Juan"
            ws.cell(row=2, column=3).value = "juan@email.com"
            ws.cell(row=2, column=4).value = ""
            ws.cell(row=2, column=5).value = "C:\\Path\\To\\Attachment.pdf"
            for col_letter, width in zip(["A", "B", "C", "D", "E"], [15, 25, 30, 30, 40]):
                ws.column_dimensions[col_letter].width = width
            wb.save(file_path)
            self.show_toast("Excel template exported successfully.", "success")
        except Exception as e:
            self.show_toast(f"Export failed: {e}", "error")

    # =================================================
    # LOAD EXCEL
    # =================================================
    def load_excel(self):
        # Use last path from settings if available
        last_path = self.settings.get("last_excel_path", "")
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Excel", last_path, "Excel Files (*.xlsx)")
        if not file_path:
            return
        try:
            df = pd.read_excel(file_path, sheet_name="Email Sending Setup")
            df = df.rename(columns={
                df.columns[0]: "Account",
                df.columns[1]: "Full Name",
                df.columns[2]: "Email",
                df.columns[3]: "CC",
                df.columns[4]: "Attachment Path"
            })
            df = df.fillna("")
            
            # Validate emails
            valid_df, invalid_emails = validate_emails_in_dataframe(df, "Email")
            
            # Show validation results
            if invalid_emails:
                error_msg = "The following emails have invalid format:\n\n"
                for item in invalid_emails[:10]:  # Show max 10 errors
                    error_msg += f"Row {item['row']}: {item['email']} - {item['error']}\n"
                
                if len(invalid_emails) > 10:
                    error_msg += f"... and {len(invalid_emails) - 10} more errors\n"
                
                error_msg += "\nThese rows will be excluded from sending."
                QMessageBox.warning(self, "Email Validation Warning", error_msg)
            
            # Use validated dataframe
            df = valid_df
            df["Status"] = EmailState.PENDING
            self.df = df[["Account", "Full Name", "Email", "CC", "Attachment Path", "Status"]]
            self.populate_table()
            
            # Update recipient counter
            recipient_count = len(self.df)
            self.recipient_counter.setText(f"{recipient_count} recipient{'s' if recipient_count != 1 else ''}")
            
            # Save the path for next time
            self.settings.set("last_excel_path", file_path)

            # Reset progress for the new import
            self.reset_progress_bars(recipient_count)
            
            if len(invalid_emails) > 0:
                self.log(f"Excel loaded with {len(invalid_emails)} invalid emails excluded.")
                self.show_toast(f"Loaded {recipient_count} recipients ({len(invalid_emails)} invalid excluded).", "warning")
            else:
                self.log("Excel loaded and validated successfully.")
                self.show_toast(f"Successfully imported {recipient_count} recipients.", "success")
            self.update_ui_state()
            self.refresh_dashboard()
        except Exception as e:
            self.show_toast(f"Failed to load Excel: {e}", "error")

    # =================================================
    # POPULATE TABLE
    # =================================================
    def populate_table(self):
        self.table.setRowCount(len(self.df))
        self.table.setColumnCount(len(self.df.columns))
        self.table.setHorizontalHeaderLabels(self.df.columns)
        
        # Set column widths
        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)  # Don't stretch last section (Status)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Account
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Full Name
        header.setSectionResizeMode(2, QHeaderView.Stretch)  # Email
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # CC
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Attachment Path
        header.setSectionResizeMode(5, QHeaderView.Fixed)  # Status - fixed width
        header.setDefaultSectionSize(100)  # Default width for stretch columns
        
        # Set specific width for Status column
        header.resizeSection(5, 100)  # Status column width
        
        for i in range(len(self.df)):
            for j in range(len(self.df.columns)):
                item = QTableWidgetItem(str(self.df.iloc[i, j]))
                
                # Color code status
                if j == 5:  # Status column
                    status = str(self.df.iloc[i, j]).lower()
                    if status == "confirmed" or status == "sent":
                        item.setBackground(QColor("#d4edda"))
                    elif status == "failed":
                        item.setBackground(QColor("#f8d7da"))
                    elif status == "unknown":
                        item.setBackground(QColor("#fff3cd"))
                    elif status == "pending":
                        item.setBackground(QColor("#e2e8f0"))
                    elif status == "sending":
                        item.setBackground(QColor("#cce5ff"))
                
                self.table.setItem(i, j, item)

    # =================================================
    # START / STOP SENDING
    # =================================================
    def start_sending(self):
        if self.df is None:
            QMessageBox.warning(self, "⚠️ Warning", "Please load an Excel file first.")
            return
        if len(self.df) == 0:
            QMessageBox.warning(self, "⚠️ Warning", "No rows to send.")
            return

        subject = self.subject_input.text()
        body = self.email_editor.toHtml()

        # Get settings
        try:
            spacing_px = int(self.spacing_select.currentData())
            max_retries = self.settings.get("max_retries", 3)
            send_delay = self.settings.get("send_delay", 0)
            importance = self.settings.get("importance", 2)
            request_read_receipt = self.settings.get("request_read_receipt", True)
        except Exception:
            spacing_px = 12
            max_retries = 3
            send_delay = 0
            importance = 2
            request_read_receipt = True

        # Save current settings
        self.settings.set("paragraph_spacing", spacing_px)

        # Create worker with enhanced safety features
        self.worker = EmailWorker(
            self.df, subject, body, spacing_px, max_retries,
            send_delay=send_delay,
            importance=importance,
            request_read_receipt=request_read_receipt,
            campaign_name=f"Campaign {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            db_manager=self.db_manager
        )
        
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.log_updated.connect(self.log)
        self.worker.status_updated.connect(self.update_status)
        self.worker.validation_complete.connect(self.on_validation_complete)
        self.worker.campaign_created.connect(self.on_campaign_created)
        self.worker.finished_sending.connect(self.finish_message)
        self.worker.start()
        
        # Update UI state
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.pause_button.setEnabled(True)
        self.resume_button.setEnabled(False)
        self.stats_timer.start(1000)
        self.progress_bar.setFormat("Sending...")
        self.dashboard_progress.setFormat("Sending...")
        self.navigate_to("logs")
        self.log("Email sending process started.")

    def stop_sending(self):
        if self.worker:
            self.worker.stop()
            self.stats_timer.stop()  # Stop statistics update timer
            self.log("⛔ Sending stopped by user.")
            self.update_ui_state()
    
    def pause_sending(self):
        if self.worker:
            self.worker.pause()
            self.log("⏸️ Sending paused.")
            self.pause_button.setEnabled(False)
            self.resume_button.setEnabled(True)
    
    def resume_sending(self):
        if self.worker:
            self.worker.resume()
            self.log("▶️ Sending resumed.")
            self.pause_button.setEnabled(True)
            self.resume_button.setEnabled(False)
    
    def retry_failed_emails(self):
        """Retry only failed emails"""
        if self.df is None:
            QMessageBox.warning(self, "⚠️ Warning", "Please load an Excel file first.")
            return
        
        # Find failed emails
        failed_indices = []
        for idx in range(len(self.df)):
            status = str(self.df.iloc[idx, 5]).lower()
            if status == "failed":
                failed_indices.append(idx)
        
        if not failed_indices:
            QMessageBox.information(self, "No Failed Emails", "There are no failed emails to retry.")
            return
        
        reply = QMessageBox.question(self, "Confirm Retry", 
                                   f"Retry {len(failed_indices)} failed emails?",
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            # Reset failed status to pending
            for idx in failed_indices:
                self.df.iloc[idx, 5] = EmailState.PENDING
                self.table.setItem(idx, 5, QTableWidgetItem(EmailState.PENDING))
            
            self.log(f"🔄 {len(failed_indices)} failed emails reset to pending. Click Start Sending to retry.")
            self.update_ui_state()
    
    def test_send(self):
        """Send a test email to verify configuration"""
        if self.df is None or len(self.df) == 0:
            QMessageBox.warning(self, "⚠️ Warning", "Please load an Excel file first.")
            return
        
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QDialogButtonBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Test Send")
        dialog.setMinimumSize(400, 200)
        dialog.setStyleSheet(build_stylesheet(self._theme))
        
        layout = QVBoxLayout(dialog)
        
        # Recipient selection
        layout.addWidget(QLabel("Select recipient for test email:"))
        recipient_combo = QComboBox()
        
        for idx in range(len(self.df)):
            name = str(self.df.iloc[idx, 1])
            email = str(self.df.iloc[idx, 2])
            recipient_combo.addItem(f"{name} ({email})", idx)
        
        layout.addWidget(recipient_combo)
        
        # Dialog buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        if dialog.exec() == QDialog.Accepted:
            idx = recipient_combo.currentData()
            row = self.df.iloc[idx]
            
            subject = self.subject_input.text()
            body = self.email_editor.toHtml()
            spacing_px = int(self.spacing_select.currentData())
            
            # Create a single-item worker for test send
            test_df = self.df.iloc[[idx]].copy()
            
            self.worker = EmailWorker(
                test_df, subject, body, spacing_px, 0,
                send_delay=0,
                importance=self.settings.get("importance", 2),
                request_read_receipt=self.settings.get("request_read_receipt", True),
                campaign_name=f"Test Send {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                db_manager=None,  # Don't save test sends to database
                is_test_send=True
            )
            
            self.worker.progress_updated.connect(self.update_progress)
            self.worker.log_updated.connect(self.log)
            self.worker.status_updated.connect(self.update_status)
            self.worker.validation_complete.connect(self.on_validation_complete)
            self.worker.finished_sending.connect(lambda: self.finish_test_send(dialog))
            self.worker.start()
            self.navigate_to("logs")
            self.log("Test send initiated...")
    
    def finish_test_send(self, dialog):
        """Handle test send completion"""
        self.show_toast("Test email has been sent.", "success")
        self.update_ui_state()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, 'toast'):
            self.toast._reposition()
    
    def view_campaign_history(self):
        self.navigate_to("history")
        self.refresh_history_page()
    
    def view_campaign_details(self, parent_dialog, campaign_list):
        """View detailed information about a selected campaign"""
        selected_items = campaign_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(parent_dialog, "No Selection", "Please select a campaign to view.")
            return
        
        campaign_id = selected_items[0].data(Qt.UserRole)
        
        try:
            recipients = self.db_manager.get_campaign_recipients(campaign_id)
            if recipients:
                from PySide6.QtWidgets import QDialog, QVBoxLayout, QTextEdit
                
                details_dialog = QDialog(parent_dialog)
                details_dialog.setWindowTitle(f"Campaign Details: {campaign_id}")
                details_dialog.setMinimumSize(900, 700)
                details_dialog.setStyleSheet(build_stylesheet(self._theme))
                
                layout = QVBoxLayout(details_dialog)
                
                details_text = QTextEdit()
                details_text.setReadOnly(True)
                
                # Build details text
                details = f"Campaign ID: {campaign_id}\n\n"
                details += "RECIPIENTS:\n"
                details += "=" * 80 + "\n"
                
                for rec in recipients:
                    (rec_id, account, full_name, email, cc, attachment_path, status, 
                     attempt_number, last_error, row_index) = rec
                    details += f"Row {row_index + 2}: {full_name} ({email})\n"
                    details += f"  Status: {status} | Attempts: {attempt_number}\n"
                    if last_error:
                        details += f"  Last Error: {last_error}\n"
                    details += "\n"
                
                details_text.setPlainText(details)
                layout.addWidget(details_text)
                
                close_button = QPushButton("Close")
                close_button.clicked.connect(details_dialog.close)
                layout.addWidget(close_button)
                
                details_dialog.exec()
            else:
                QMessageBox.warning(parent_dialog, "Error", "No recipients found for this campaign.")
        except Exception as e:
            QMessageBox.critical(parent_dialog, "Error", f"Failed to load campaign details: {str(e)}")
    
    def delete_campaign(self, parent_dialog, campaign_list):
        """Delete a selected campaign"""
        selected_items = campaign_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(parent_dialog, "No Selection", "Please select a campaign to delete.")
            return
        
        campaign_id = selected_items[0].data(Qt.UserRole)
        
        reply = QMessageBox.question(parent_dialog, "Confirm Delete", 
                                   f"Are you sure you want to delete campaign {campaign_id}?\n\nThis action cannot be undone.",
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                conn = self.db_manager.get_connection()
                cursor = conn.cursor()
                
                # Delete related records first
                cursor.execute('DELETE FROM send_logs WHERE campaign_id = ?', (campaign_id,))
                cursor.execute('DELETE FROM send_attempts WHERE campaign_id = ?', (campaign_id,))
                cursor.execute('DELETE FROM recipients WHERE campaign_id = ?', (campaign_id,))
                cursor.execute('DELETE FROM campaigns WHERE id = ?', (campaign_id,))
                
                conn.commit()
                conn.close()
                
                # Refresh the list
                parent_dialog.close()
                self.view_campaign_history()  # Reopen with updated data
                
                QMessageBox.information(self, "Success", f"Campaign {campaign_id} has been deleted.")
            except Exception as e:
                QMessageBox.critical(parent_dialog, "Error", f"Failed to delete campaign: {str(e)}")

    def on_campaign_created(self, campaign_id):
        """Handle campaign creation"""
        self.current_campaign_id = campaign_id
        self.log(f"📋 Campaign ID: {campaign_id}")
    
    def on_validation_complete(self, is_valid, validation_results):
        """Handle validation completion"""
        if not is_valid and not (self.worker and self.worker.is_test_send):
            # Show validation error dialog (skip for test sends)
            error_msg = "Excel Validation Failed\n\n"
            
            if validation_results['invalid_emails']:
                error_msg += f"Invalid Emails ({len(validation_results['invalid_emails'])}):\n"
                for item in validation_results['invalid_emails'][:5]:
                    error_msg += f"  Row {item['row']}: {item['email']} - {item['error']}\n"
                if len(validation_results['invalid_emails']) > 5:
                    error_msg += f"  ... and {len(validation_results['invalid_emails']) - 5} more\n"
                error_msg += "\n"
            
            if validation_results['missing_attachments']:
                error_msg += f"Missing Attachments ({len(validation_results['missing_attachments'])}):\n"
                for item in validation_results['missing_attachments'][:5]:
                    error_msg += f"  Row {item['row']}: {item['path']}\n"
                if len(validation_results['missing_attachments']) > 5:
                    error_msg += f"  ... and {len(validation_results['missing_attachments']) - 5} more\n"
                error_msg += "\n"
            
            if validation_results['empty_required_fields']:
                error_msg += f"Empty Required Fields ({len(validation_results['empty_required_fields'])}):\n"
                for item in validation_results['empty_required_fields'][:5]:
                    error_msg += f"  Row {item['row']}: {item['field']}\n"
                if len(validation_results['empty_required_fields']) > 5:
                    error_msg += f"  ... and {len(validation_results['empty_required_fields']) - 5} more\n"
                error_msg += "\n"
            
            if validation_results['invalid_cc']:
                error_msg += f"Invalid CC Addresses ({len(validation_results['invalid_cc'])}):\n"
                for item in validation_results['invalid_cc'][:5]:
                    error_msg += f"  Row {item['row']}: {item['cc']} - {item['error']}\n"
                if len(validation_results['invalid_cc']) > 5:
                    error_msg += f"  ... and {len(validation_results['invalid_cc']) - 5} more\n"
                error_msg += "\n"
            
            if validation_results['duplicate_recipients']:
                error_msg += f"Duplicate Recipients ({len(validation_results['duplicate_recipients'])}):\n"
                for item in validation_results['duplicate_recipients'][:5]:
                    error_msg += f"  Row {item['row']}: {item['email']} (duplicate of row {item['duplicate_row']})\n"
                if len(validation_results['duplicate_recipients']) > 5:
                    error_msg += f"  ... and {len(validation_results['duplicate_recipients']) - 5} more\n"
                error_msg += "\n"
            
            error_msg += "Please fix these issues before sending."
            QMessageBox.critical(self, "Validation Failed", error_msg)

    # =================================================
    # UPDATE STATUS & LOGS
    # =================================================
    def reset_progress_bars(self, recipient_count=0):
        """Reset sending progress UI after a new import or before a new campaign."""
        self.progress_bar.setValue(0)
        self.dashboard_progress.setValue(0)
        if recipient_count > 0:
            label = f"Ready — {recipient_count} recipient{'s' if recipient_count != 1 else ''} loaded"
        else:
            label = "Ready — load recipients to begin"
        self.progress_bar.setFormat(label)
        self.dashboard_progress.setFormat(label)

    def update_progress(self, value: int):
        self.progress_bar.setValue(value)
        self.progress_bar.setFormat(f"{value}% complete")
        self.dashboard_progress.setValue(value)
        self.dashboard_progress.setFormat(f"{value}% complete")

    def update_status(self, row, status):
        self.table.setItem(row, 5, QTableWidgetItem(status))
        # Update the dataframe as well
        if self.df is not None and row < len(self.df):
            self.df.iloc[row, 5] = status
    
    def update_statistics(self):
        """Update statistics cards"""
        if self.df is not None and len(self.df) > 0:
            recipient_count = len(self.df)
            confirmed = sum(1 for idx in range(len(self.df)) if str(self.df.iloc[idx, 5]).lower() == "confirmed")
            failed = sum(1 for idx in range(len(self.df)) if str(self.df.iloc[idx, 5]).lower() == "failed")
            pending = sum(1 for idx in range(len(self.df)) if str(self.df.iloc[idx, 5]).lower() == "pending")
            unknown = sum(1 for idx in range(len(self.df)) if str(self.df.iloc[idx, 5]).lower() == "unknown")
            
            # Update stat cards
            self.stat_total_value.setText(str(recipient_count))
            self.stat_sent_value.setText(str(confirmed))
            self.stat_failed_value.setText(str(failed))
            self.stat_pending_value.setText(str(pending))
            
            # Update recipient counter with status breakdown
            counter_text = f"{recipient_count} recipient{'s' if recipient_count != 1 else ''}"
            if confirmed > 0 or failed > 0 or pending > 0 or unknown > 0:
                counter_text += f"  ·  Sent {confirmed}  ·  Failed {failed}  ·  Pending {pending}"
            self.recipient_counter.setText(counter_text)

    def log(self, message):
        """Log message to both UI and file"""
        self.log_box.append(message)
        
        # Also log to file
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        
        try:
            with open("email_sender.log", "a", encoding="utf-8") as f:
                f.write(log_message + "\n")
        except Exception:
            pass  # Silently fail if logging doesn't work

    # =================================================
    # LOGGING SYSTEM
    # =================================================
    def setup_logging(self):
        """Setup file-based logging system"""
        try:
            # Create logs directory if it doesn't exist
            if not os.path.exists("logs"):
                os.makedirs("logs")
            
            # Setup logging configuration
            log_file = os.path.join("logs", f"email_sender_{datetime.now().strftime('%Y%m%d')}.log")
            
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(log_file, encoding='utf-8'),
                    logging.StreamHandler()  # Also log to console
                ]
            )
            
            # Log application start
            logging.info("Eru Email Sender Pro started")
            
        except Exception as e:
            print(f"Failed to setup logging: {e}")

    def finish_message(self):
        self.stats_timer.stop()
        self.progress_bar.setFormat("Complete")
        self.dashboard_progress.setFormat("Complete")
        self.refresh_dashboard()
        self.show_toast("Email sending process has completed.", "success")
        self.update_ui_state()

    # =================================================
    # TEXT FORMATTING
    # =================================================
    def make_bold(self):
        fmt = self.email_editor.currentCharFormat()
        fmt.setFontWeight(QFont.Bold if fmt.fontWeight() != QFont.Bold else QFont.Normal)
        self.email_editor.setCurrentCharFormat(fmt)

    def make_italic(self):
        fmt = self.email_editor.currentCharFormat()
        fmt.setFontItalic(not fmt.fontItalic())
        self.email_editor.setCurrentCharFormat(fmt)

    def make_underline(self):
        fmt = self.email_editor.currentCharFormat()
        fmt.setFontUnderline(not fmt.fontUnderline())
        self.email_editor.setCurrentCharFormat(fmt)

    # =================================================
    # EMAIL PREVIEW
    # =================================================
    def preview_email(self):
        """Preview email with sample data"""
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QTextBrowser, QHBoxLayout, QLabel, QComboBox
        
        # Create preview dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Email Preview")
        dialog.setMinimumSize(800, 600)
        dialog.setStyleSheet(build_stylesheet(self._theme))
        
        layout = QVBoxLayout(dialog)
        
        # Sample recipient selection
        sample_row = QHBoxLayout()
        sample_label = QLabel("Sample Recipient:")
        sample_combo = QComboBox()
        sample_combo.addItem("Sample: ACC-001 - Dela Cruz, Juan", {"account": "ACC-001", "name": "Dela Cruz, Juan"})
        sample_combo.addItem("Sample: ACC-002 - Smith, John", {"account": "ACC-002", "name": "Smith, John"})
        sample_combo.addItem("Sample: ACC-003 - Garcia, Maria", {"account": "ACC-003", "name": "Garcia, Maria"})
        sample_row.addWidget(sample_label)
        sample_row.addWidget(sample_combo)
        sample_row.addStretch()
        layout.addLayout(sample_row)
        
        # Preview browser
        preview_browser = QTextBrowser()
        preview_browser.setObjectName("previewBrowser")
        layout.addWidget(preview_browser)
        
        # Update preview when sample changes
        def update_preview():
            sample_data = sample_combo.currentData()
            sample_account = sample_data["account"]
            sample_name = sample_data["name"]
            subject = self.subject_input.text().replace("{{account}}", sample_account).replace("{{Account}}", sample_account).replace("{{fullname}}", sample_name)
            body_html = self.email_editor.toHtml().replace("{{fullname}}", get_surname(sample_name))
            
            # Apply Outlook-safe formatting
            try:
                spacing_px = int(self.spacing_select.currentData())
                final_html = build_outlook_safe_html(body_html, spacing_px)
            except Exception:
                final_html = body_html
            
            preview_html = f"""
            <div style="font-family: Segoe UI, Arial, sans-serif; padding: 20px;">
                <h3 style="color: #1e40af; margin-bottom: 15px;">Subject: {subject}</h3>
                <div style="border-top: 1px solid #e2e8f0; padding-top: 15px;">
                    {final_html}
                </div>
            </div>
            """
            preview_browser.setHtml(preview_html)
        
        sample_combo.currentIndexChanged.connect(lambda: update_preview())
        update_preview()  # Initial preview
        
        # Dialog buttons
        from PySide6.QtWidgets import QDialogButtonBox
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(dialog.close)
        layout.addWidget(buttons)
        
        dialog.exec()

    # =================================================
    # COMPOSER SPACING PREVIEW
    # =================================================
    def apply_editor_paragraph_spacing(self, px: int):
        doc = self.email_editor.document()
        cursor = QTextCursor(doc)
        cursor.beginEditBlock()
        block = doc.begin()
        while block.isValid():
            bfmt = block.blockFormat()
            bfmt.setTopMargin(0)
            bfmt.setBottomMargin(max(0, int(px)))
            # Use proportional line height ~135% for readability
            try:
                bfmt.setLineHeight(135, QTextBlockFormat.ProportionalHeight)
            except Exception:
                pass
            c = QTextCursor(block)
            c.setBlockFormat(bfmt)
            block = block.next()
        cursor.endEditBlock()

    def on_spacing_changed(self, _index: int):
        try:
            px = int(self.spacing_select.currentData())
        except Exception:
            px = 12
        self.apply_editor_paragraph_spacing(px)
        self.settings.set("paragraph_spacing", px)

    # =================================================
    # TEMPLATE MANAGEMENT
    # =================================================
    def load_templates(self):
        """Load templates from settings into combo box"""
        templates = self.settings.get("email_templates", {})
        
        # Clear existing items except default
        self.template_combo.clear()
        self.template_combo.addItem("Default HR Notice", "default")
        
        # Add saved templates
        for name in templates.keys():
            self.template_combo.addItem(name, name)
        
        # Restore last selected template
        last_template = self.settings.get("last_selected_template", "default")
        
        template_found = False
        for i in range(self.template_combo.count()):
            item_data = self.template_combo.itemData(i)
            if item_data == last_template:
                # Temporarily block signals to avoid triggering load_template twice
                self.template_combo.blockSignals(True)
                self.template_combo.setCurrentIndex(i)
                self.template_combo.blockSignals(False)
                # Load the template content directly without signal
                self._load_template_content(i)
                template_found = True
                break
        
        if not template_found:
            # Default template is already selected at index 0
            self._load_template_content(0)
    
    def _load_template_content(self, index):
        """Load template content without saving the last selected template"""
        if index == 0:  # Default template
            self.subject_input.setText("NOTICE TO SUBMIT LACKING EMPLOYMENT REQUIREMENTS - {{fullname}} - {{account}}")
            default_body = """
<p>Dear {{fullname}},</p>

<p>This is to formally inform you that you still have outstanding mandatory employment requirements as of this date, despite prior reminders and your signed Affidavit of Undertaking upon commencement of employment.</p>

<p>As stated in your Affidavit of Undertaking, you committed to submit all required documents within prescribed period. <b>You are hereby given five (5) days from receipt of this email notice </b> to complete and submit pending requirements. Please see the attached <b>Notice of Incomplete Employment Requirements</b> for full details. Failure to comply within the given timeframe, may result in appropriate administrative action in accordance with Company policy.</p>

<p>Please submit the required documents through this same email thread. For any clarification, please coordinate with <b>HR-DMRC or your assigned account supervisor.</b></p>

<p>Thanks,<br>Jhudel S. Orola<br>HR Staff - Data Management & Records Control<br>Acabar Marketing International Inc.<br>(02) 8887-8170 Local 153</p>

<p><img class="x_CToWUd" height="77" width="250" src="https://ci3.googleusercontent.com/mail-sig/AIorK4x0oCXqeBBsjR9hQB3HLxhAJPc1msod_2dqrIiATYz-sDfATgJdOa_R6eWlr16--ykbMmeApG_G3we-" data-imagetype="External"></p>
"""
            self.email_editor.setHtml(default_body)
        else:
            template_name = self.template_combo.itemData(index)  # Use index instead of currentData
            templates = self.settings.get("email_templates", {})
            if template_name in templates:
                template = templates[template_name]
                self.subject_input.setText(template.get("subject", ""))
                self.email_editor.setHtml(template.get("body", ""))
        
        # Apply current spacing
        try:
            px = int(self.spacing_select.currentData())
            self.apply_editor_paragraph_spacing(px)
        except Exception:
            pass

    def load_template(self, index):
        """Load selected template into editor"""
        # Load the content
        self._load_template_content(index)
        
        # Save the last selected template
        template_name = self.template_combo.currentData()
        self.settings.set("last_selected_template", template_name)
    
    def save_template(self):
        """Save current email as template"""
        from PySide6.QtWidgets import QInputDialog
        
        name, ok = QInputDialog.getText(self, "Save Template", "Enter template name:")
        if not ok or not name.strip():
            return
        
        name = name.strip()
        templates = self.settings.get("email_templates", {})
        templates[name] = {
            "subject": self.subject_input.text(),
            "body": self.email_editor.toHtml()
        }
        
        self.settings.set("email_templates", templates)
        self.load_templates()  # Refresh combo box
        
        # Select the newly saved template
        for i in range(self.template_combo.count()):
            if self.template_combo.itemData(i) == name:
                self.template_combo.setCurrentIndex(i)
                break
        
        QMessageBox.information(self, "Success", f"Template '{name}' saved successfully.")
        self.show_toast(f"Template '{name}' saved.", "success")
        self.refresh_templates_page()
    
    def _delete_template_from_page(self):
        row = self.template_list_widget.currentRow()
        if row <= 0:
            self.show_toast("Cannot delete the default template.", "warning")
            return
        name = self.template_list_widget.item(row).text()
        reply = QMessageBox.question(self, "Confirm Delete",
                                     f"Delete template '{name}'?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            templates = self.settings.get("email_templates", {})
            if name in templates:
                del templates[name]
                self.settings.set("email_templates", templates)
                self.load_templates()
                self.template_combo.setCurrentIndex(0)
                self.refresh_templates_page()
                self.show_toast(f"Template '{name}' deleted.", "success")

    def delete_template(self):
        """Delete selected template"""
        if self.template_combo.currentIndex() == 0:
            QMessageBox.warning(self, "Warning", "Cannot delete the default template.")
            return
        
        template_name = self.template_combo.currentData()
        reply = QMessageBox.question(self, "Confirm Delete", 
                                   f"Are you sure you want to delete template '{template_name}'?",
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            templates = self.settings.get("email_templates", {})
            if template_name in templates:
                del templates[template_name]
                self.settings.set("email_templates", templates)
                self.load_templates()  # Refresh combo box
                self.template_combo.setCurrentIndex(0)  # Select default
                QMessageBox.information(self, "Success", f"Template '{template_name}' deleted.")

    # =================================================
    # KEYBOARD SHORTCUTS
    # =================================================
    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts for common actions"""
        
        # Export Template: Ctrl+E
        export_shortcut = QAction(self)
        export_shortcut.setShortcut(QKeySequence("Ctrl+E"))
        export_shortcut.triggered.connect(self.export_template)
        self.addAction(export_shortcut)
        
        # Load Excel: Ctrl+O
        load_shortcut = QAction(self)
        load_shortcut.setShortcut(QKeySequence("Ctrl+O"))
        load_shortcut.triggered.connect(self.load_excel)
        self.addAction(load_shortcut)
        
        # Start Sending: Ctrl+S
        start_shortcut = QAction(self)
        start_shortcut.setShortcut(QKeySequence("Ctrl+S"))
        start_shortcut.triggered.connect(self.start_sending)
        self.addAction(start_shortcut)
        
        # Stop Sending: Ctrl+Shift+S
        stop_shortcut = QAction(self)
        stop_shortcut.setShortcut(QKeySequence("Ctrl+Shift+S"))
        stop_shortcut.triggered.connect(self.stop_sending)
        self.addAction(stop_shortcut)
        
        # Preview Email: Ctrl+P
        preview_shortcut = QAction(self)
        preview_shortcut.setShortcut(QKeySequence("Ctrl+P"))
        preview_shortcut.triggered.connect(self.preview_email)
        self.addAction(preview_shortcut)
        
        # Save Template: Ctrl+T
        save_template_shortcut = QAction(self)
        save_template_shortcut.setShortcut(QKeySequence("Ctrl+T"))
        save_template_shortcut.triggered.connect(self.save_template)
        self.addAction(save_template_shortcut)
        
        # Bold: Ctrl+B
        bold_shortcut = QAction(self)
        bold_shortcut.setShortcut(QKeySequence("Ctrl+B"))
        bold_shortcut.triggered.connect(self.make_bold)
        self.addAction(bold_shortcut)
        
        # Italic: Ctrl+I
        italic_shortcut = QAction(self)
        italic_shortcut.setShortcut(QKeySequence("Ctrl+I"))
        italic_shortcut.triggered.connect(self.make_italic)
        self.addAction(italic_shortcut)
        
        # Underline: Ctrl+U
        underline_shortcut = QAction(self)
        underline_shortcut.setShortcut(QKeySequence("Ctrl+U"))
        underline_shortcut.triggered.connect(self.make_underline)
        self.addAction(underline_shortcut)

    # =================================================
    # MODERN STYLES
    # =================================================
    def modern_styles(self):
        return build_stylesheet(self._theme)


# =====================================================
# RUN APPLICATION
# =====================================================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = EmailApp()
    window.show()
    sys.exit(app.exec())